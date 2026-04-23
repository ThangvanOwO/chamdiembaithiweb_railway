"""
==========================================================================
 HỆ THỐNG CHẤM BÀI TRẮC NGHIỆM OMR (Optical Mark Recognition)
 Phiếu trắc nghiệm Việt Nam - OpenCV thuần (không Deep Learning)
==========================================================================
Cấu trúc phiếu:
  - Phần I  : 40 câu ABCD, 4 cột x 10 hàng
  - Phần II : 8 câu, mỗi câu a/b/c/d x Đúng/Sai
  - Phần III: 6 câu điền số (dấu trừ, dấu phẩy, 4 cột số 0-9)

Tọa độ bubble đã calibrate bằng HoughCircles trên ảnh warped 1400x1920.

Chống nhiễu chữ in (A, B, C, D, số thứ tự...) — 3 LỚP BẢO VỆ:
  LỚP 1) erase_printed_text(): tô TRẮNG chữ in trên ảnh warped TRƯỚC threshold
         → Xóa mạnh vùng rộng nhưng BẢO VỆ bubble bằng lỗ tròn (punch-holes)
  LỚP 2) Morphological opening: loại nét mảnh còn sót (viền, stroke nhỏ)
  LỚP 3) mask_printed_text(): tô ĐEN vùng text trên threshold (safety net)
"""

import cv2
import numpy as np
import os
import json
import logging
from datetime import datetime
from PIL import Image, ExifTags
from skimage import exposure

logger = logging.getLogger(__name__)

# ╔════════════════════════════════════════════════════════════════════════╗
# ║                        CẤU HÌNH CHUNG                               ║
# ╚════════════════════════════════════════════════════════════════════════╝

# --- Kích thước ảnh warped (sau perspective transform) ---
WARP_WIDTH = 1400
WARP_HEIGHT = 1920

# --- Vùng crop tên học sinh trên ảnh warped ---
# (x, y, w, h) — dòng "4. Họ và tên thí sinh:......."
NAME_REGION = (255, 300, 745, 60)

# --- Tham số phát hiện bubble ---
# Ngưỡng tỷ lệ pixel đen / tổng pixel trong vòng tròn
# Bubble rỗng (viền) ~0.05-0.12 trên ảnh cleaned
# Bubble đã tô        ~0.45-0.85
# Chỉnh trong khoảng 0.28 - 0.45 tùy chất lượng in/scan
FILL_THRESHOLD = 0.38

# Bán kính bubble (pixel trên ảnh warped, đo từ HoughCircles ~11-14)
BUBBLE_RADIUS = 13

# Kích thước kernel morphological opening (LỚP 2) để loại nét chữ in
# Kernel 5x5 loại nét < 5px (chữ in ~2-3px, viền bubble ~2px)
# === CHỈNH KERNEL Ở ĐÂY === (5=an toàn, 7=mạnh hơn nhưng có thể phá bubble nhạt)
MORPH_KERNEL_SIZE = 3

# Ngưỡng circularity (hình tròn = 1.0, chữ cái < 0.5)
# Bubble tô đặc: circularity ~0.7-1.0
# Chữ cái A,B,C,D: circularity ~0.2-0.5
# === CHỈNH CIRCULARITY Ở ĐÂY === (0=tắt, 0.4-0.7=lọc text, >0.7=quá strict)
CIRCULARITY_THRESHOLD = 0.6

# Bán kính vùng bảo vệ bubble khi xóa text (LỚP 1)
# Phải >= BUBBLE_RADIUS để không xóa vào bubble
# +8px margin cho sai lệch warp ảnh phone (trước: +3px quá ít)
BUBBLE_PROTECT_RADIUS = BUBBLE_RADIUS + 8

# --- Hybrid OpenCV + CNN (bubble classifier) ---
HYBRID_CNN_ENABLE = True
BUBBLE_CNN_PATH = os.path.join(os.path.dirname(__file__), "bubble_cnn.pth")
CNN_IMG_SIZE = 32
CNN_FILL_THRESHOLD = 0.5
HYBRID_RATIO_LOW = 0.12
HYBRID_RATIO_HIGH = 0.45

# Part III hybrid thresholds (scores are 0-1)
P3_SIGN_SCORE_MIN = 0.18
P3_COMMA_SCORE_MIN = 0.10
P3_COMMA_GAP_MIN = 0.05
P3_DIGIT_SCORE_MIN = 0.28
P3_DIGIT_GAP_MIN = 0.05
P3_OCR_ENABLE = True
P3_OCR_BOX_Y_OFFSET = -55  # relative to PART3_SIGN_Y
P3_OCR_BOX_SIZE = 34
P3_OCR_INK_MIN = 0.04
P3_OCR_INK_MAX = 0.45

# SBD/Mã đề thresholds (lighter pencil marks)
SBD_MADE_TOP_MIN = 0.12
SBD_MADE_GAP_MIN = 0.02
SBD_MADE_FALLBACK_TOP_MIN = 0.10
SBD_MADE_FALLBACK_GAP_MIN = 0.08

_CNN_MODEL = None
_CNN_DEVICE = None
_CNN_READY = False
_CNN_ERROR = None

# ╔════════════════════════════════════════════════════════════════════════╗
# ║              TỌA ĐỘ PHẦN I - 40 câu ABCD (4 cột x 10 hàng)        ║
# ╚════════════════════════════════════════════════════════════════════════╝
# start_x, start_y: tâm bubble đầu tiên (lựa chọn A, câu 1 của cột)
# step_x: khoảng cách ngang giữa A→B→C→D (~73px)
# step_y: khoảng cách dọc giữa các câu (~33px)
PART1_COLS = [
    {"start_x": 82,   "start_y": 689, "step_x": 72.3, "step_y": 33.1, "q_start": 1},   # Cột 1: Q1-Q10
    {"start_x": 430,  "start_y": 691, "step_x": 74,   "step_y": 33.1, "q_start": 11},  # Cột 2: Q11-Q20
    {"start_x": 781,  "start_y": 691, "step_x": 73,   "step_y": 33.1, "q_start": 21},  # Cột 3: Q21-Q30
    {"start_x": 1130, "start_y": 691, "step_x": 73,   "step_y": 33.1, "q_start": 31},  # Cột 4: Q31-Q40
]
PART1_NUM_ROWS = 10
PART1_CHOICES = ["A", "B", "C", "D"]

# ╔════════════════════════════════════════════════════════════════════════╗
# ║       TỌA ĐỘ PHẦN II - 8 câu x (a/b/c/d) x (Đúng/Sai)           ║
# ╚════════════════════════════════════════════════════════════════════════╝
# Mỗi block: start_x = tâm cột Đúng, start_x + step = tâm cột Sai
PART2_BLOCKS = [
    {"start_x": 81,   "start_y": 1187, "q": 1},
    {"start_x": 228,  "start_y": 1187, "q": 2},
    {"start_x": 430,  "start_y": 1187, "q": 3},
    {"start_x": 577,  "start_y": 1187, "q": 4},
    {"start_x": 781,  "start_y": 1187, "q": 5},
    {"start_x": 927,  "start_y": 1187, "q": 6},
    {"start_x": 1130, "start_y": 1187, "q": 7},
    {"start_x": 1276, "start_y": 1187, "q": 8},
]
PART2_STEP_X = 73   # Khoảng cách Đúng → Sai
PART2_STEP_Y = 33   # Khoảng cách a → b → c → d
PART2_ROWS = ["a", "b", "c", "d"]

# ╔════════════════════════════════════════════════════════════════════════╗
# ║      TỌA ĐỘ PHẦN III - 6 câu điền số (dấu trừ, phẩy, 4 cột 0-9) ║
# ╚════════════════════════════════════════════════════════════════════════╝
# Mỗi câu: sign_x = tâm bubble dấu trừ, cols_x = [4 tâm cột số]
PART3_BLOCKS = [
    {"sign_x": 81,   "cols_x": [90,  124, 159, 192],  "q": 1},
    {"sign_x": 313,  "cols_x": [324, 357, 391, 425],  "q": 2},
    {"sign_x": 547,  "cols_x": [557, 591, 624, 659],  "q": 3},
    {"sign_x": 780,  "cols_x": [790, 823, 858, 892],  "q": 4},
    {"sign_x": 1013, "cols_x": [1023, 1057, 1091, 1125], "q": 5},
    {"sign_x": 1247, "cols_x": [1257, 1290, 1324, 1359], "q": 6},
]
PART3_SIGN_Y = 1490         # Hàng dấu trừ (-)
PART3_COMMA_Y = 1522        # Hàng dấu phẩy (.)
PART3_DIGIT_START_Y = 1555  # Hàng chữ số 0
PART3_DIGIT_STEP_Y = 33.1   # Bước giữa các hàng 0→1→...→9
PART3_NUM_DIGIT_COLS = 4

# ╔════════════════════════════════════════════════════════════════════════╗
# ║     TỌA ĐỘ SỐ BÁO DANH (6 chữ số) + MÃ ĐỀ (3 chữ số)            ║
# ╚════════════════════════════════════════════════════════════════════════╝
# Vùng góc trên bên phải, mỗi cột có 10 bubble (digits 0-9)
# Tọa độ x cho từng cột (calibrated bằng HoughCircles)
SBD_COLS_X = [1057, 1085, 1113, 1141, 1169, 1197]  # 6 cột Số báo danh
MADE_COLS_X = [1292, 1321, 1345]                     # 3 cột Mã đề
# Tọa độ y cho hàng digit 0-9 (không đều nên dùng mảng tường minh)
SBD_MADE_DIGIT_Y = [173, 206, 249, 285, 326, 363, 401, 440, 480, 517]

# ╔════════════════════════════════════════════════════════════════════════╗
# ║    VÙNG CHỮ IN CẦN XÓA (aggressive — dùng với bubble protection)  ║
# ╚════════════════════════════════════════════════════════════════════════╝
# Mỗi vùng (x1, y1, x2, y2) sẽ bị xóa trắng/đen, NGOẠI TRỪ vùng tròn
# quanh mỗi bubble center (punch-holes). Nhờ đó có thể phủ RẤT RỘNG mà
# không phá hủy dữ liệu bubble.

def _build_text_erase_regions():
    """
    Tạo danh sách vùng chữ in cần xóa — RỘNG TỐI ĐA.
    Bubble được bảo vệ bởi punch-holes nên không sợ mất dữ liệu.
    """
    # === CHỈNH VÙNG MASK Ở ĐÂY ===
    regions = []

    # --- Phần I: Header "A B C D" + banner "PHẦN I" ---
    # Phủ TOÀN BỘ từ banner đến sát row 1 (y=590→680)
    # Chữ A,B,C,D thực tế ở y=655-668 → phải phủ đến y≥670
    # Bubble row 1 ở y=660 nhưng ĐƯỢC BẢO VỆ bởi punch-holes
    for cfg in PART1_COLS:
        sx = cfg["start_x"]
        ex = sx + 3 * cfg["step_x"]
        regions.append((sx - 30, 590, ex + 30, 680))
        # Số thứ tự câu bên trái: "1", "2", ..., "10"
        y_top = int(cfg["start_y"] - 15)
        col_rows = cfg.get("num_rows", PART1_NUM_ROWS)
        y_bot = int(cfg["start_y"] + (col_rows - 1) * cfg["step_y"] + 22)
        regions.append((sx - 60, y_top, sx - 3, y_bot))

    # Banner "PHẦN I" phía trên
    regions.append((15, 570, 1385, 595))

    # --- Phần II: Header "Câu N" + "Đúng Sai" + label "a) b) c) d)" ---
    regions.append((25, 1095, 1375, 1188))
    for blk in PART2_BLOCKS:
        sx = blk["start_x"]
        y_top = blk["start_y"] - 12
        y_bot = int(blk["start_y"] + 3 * PART2_STEP_Y + 12)
        regions.append((sx - 35, y_top, sx - 3, y_bot))

    # --- Phần III: Header "Câu N" + label "0"-"9" ---
    # CHÚ Ý: KHÔNG mask sign(-) và comma(.) — chỉ mask label số
    regions.append((25, 1305, 1375, 1400))
    for blk in PART3_BLOCKS:
        first_col_x = blk["cols_x"][0]
        y_top = int(PART3_DIGIT_START_Y - 8)
        y_bot = int(PART3_DIGIT_START_Y + 9 * PART3_DIGIT_STEP_Y + 15)
        regions.append((first_col_x - 45, y_top, first_col_x - 3, y_bot))

    # --- SBD + Mã đề: Label "0"-"9" bên trái ---
    sbd_left = min(SBD_COLS_X) - 45
    sbd_right = min(SBD_COLS_X) - 3
    regions.append((sbd_left, SBD_MADE_DIGIT_Y[0] - 8, sbd_right, SBD_MADE_DIGIT_Y[-1] + 15))
    made_left = min(MADE_COLS_X) - 45
    made_right = min(MADE_COLS_X) - 3
    regions.append((made_left, SBD_MADE_DIGIT_Y[0] - 8, made_right, SBD_MADE_DIGIT_Y[-1] + 15))

    return regions


TEXT_ERASE_REGIONS = _build_text_erase_regions()


def _collect_all_bubble_centers():
    """Thu thập TẤT CẢ tâm bubble trên phiếu → dùng cho punch-holes bảo vệ."""
    centers = []
    # Part I: mỗi cột có thể có num_rows riêng (hỗ trợ biến thể 26, 30 câu)
    for cfg in PART1_COLS:
        col_rows = cfg.get("num_rows", PART1_NUM_ROWS)
        for ci in range(4):
            for ri in range(col_rows):
                cx = int(cfg["start_x"] + ci * cfg["step_x"])
                cy = int(cfg["start_y"] + ri * cfg["step_y"])
                centers.append((cx, cy))
    # Part II: 8 block × 4 hàng × 2 cột = 64 bubble
    for blk in PART2_BLOCKS:
        for ci in range(2):
            for ri in range(4):
                cx = int(blk["start_x"] + ci * PART2_STEP_X)
                cy = int(blk["start_y"] + ri * PART2_STEP_Y)
                centers.append((cx, cy))
    # Part III: 6 câu × (1 sign + 4 comma + 4×10 digits)
    for blk in PART3_BLOCKS:
        centers.append((int(blk["sign_x"]), PART3_SIGN_Y))
        for col_x in blk["cols_x"]:
            centers.append((int(col_x), PART3_COMMA_Y))
            for d in range(10):
                cy = int(PART3_DIGIT_START_Y + d * PART3_DIGIT_STEP_Y)
                centers.append((int(col_x), cy))
    # SBD + Mã đề
    for cx in list(SBD_COLS_X) + list(MADE_COLS_X):
        for cy in SBD_MADE_DIGIT_Y:
            centers.append((int(cx), int(cy)))
    return centers


ALL_BUBBLE_CENTERS = _collect_all_bubble_centers()


def _build_erase_mask():
    """
    Pre-compute erase mask 1 lần duy nhất khi load module.
    Pixel=255 → sẽ bị xóa (trắng trên warped, đen trên threshold).
    Pixel=0   → được giữ nguyên (bao gồm vùng bubble).
    """
    mask = np.zeros((WARP_HEIGHT, WARP_WIDTH), dtype=np.uint8)
    # Phủ tất cả vùng text
    for (x1, y1, x2, y2) in TEXT_ERASE_REGIONS:
        x1c = max(0, int(x1))
        y1c = max(0, int(y1))
        x2c = min(WARP_WIDTH, int(x2))
        y2c = min(WARP_HEIGHT, int(y2))
        mask[y1c:y2c, x1c:x2c] = 255
    # Đục lỗ tròn (punch-holes) tại mỗi bubble center → bảo vệ bubble
    for (cx, cy) in ALL_BUBBLE_CENTERS:
        cv2.circle(mask, (cx, cy), BUBBLE_PROTECT_RADIUS, 0, -1)
    return mask


ERASE_MASK = _build_erase_mask()


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                 JSON: LOAD TEMPLATE / ANSWERS / SAVE                 ║
# ╚════════════════════════════════════════════════════════════════════════╝

def load_template(json_path):
    """
    Đọc template JSON → cập nhật toàn bộ global constants.
    Gọi hàm này TRƯỚC khi chấm bài nếu muốn dùng phiếu khác.
    """
    global WARP_WIDTH, WARP_HEIGHT, NAME_REGION
    global FILL_THRESHOLD, BUBBLE_RADIUS, MORPH_KERNEL_SIZE
    global CIRCULARITY_THRESHOLD, BUBBLE_PROTECT_RADIUS
    global PART1_COLS, PART1_NUM_ROWS, PART1_CHOICES
    global PART2_BLOCKS, PART2_STEP_X, PART2_STEP_Y, PART2_ROWS
    global PART3_BLOCKS, PART3_SIGN_Y, PART3_COMMA_Y
    global PART3_DIGIT_START_Y, PART3_DIGIT_STEP_Y, PART3_NUM_DIGIT_COLS
    global SBD_COLS_X, MADE_COLS_X, SBD_MADE_DIGIT_Y
    global TEXT_ERASE_REGIONS, ALL_BUBBLE_CENTERS, ERASE_MASK

    with open(json_path, "r", encoding="utf-8") as f:
        t = json.load(f)

    # Warp
    WARP_WIDTH  = t["warp"]["width"]
    WARP_HEIGHT = t["warp"]["height"]

    # Name region (optional override)
    if "name_region" in t:
        nr = t["name_region"]
        NAME_REGION = (nr["x"], nr["y"], nr["w"], nr["h"])

    # Detection
    det = t["detection"]
    FILL_THRESHOLD         = det["fill_threshold"]
    BUBBLE_RADIUS          = det["bubble_radius"]
    MORPH_KERNEL_SIZE      = det["morph_kernel_size"]
    CIRCULARITY_THRESHOLD  = det["circularity_threshold"]
    BUBBLE_PROTECT_RADIUS  = det.get("bubble_protect_radius", BUBBLE_RADIUS + 3)

    # Part I
    p1 = t["part1"]
    PART1_COLS     = p1["columns"]
    PART1_NUM_ROWS = p1.get("num_rows", 10)
    PART1_CHOICES  = p1.get("choices", ["A", "B", "C", "D"])

    # Part II
    p2 = t["part2"]
    PART2_BLOCKS = p2["blocks"]
    PART2_STEP_X = p2["step_x"]
    PART2_STEP_Y = p2["step_y"]
    PART2_ROWS   = p2.get("rows", ["a", "b", "c", "d"])

    # Part III
    p3 = t["part3"]
    PART3_BLOCKS         = p3["blocks"]
    PART3_SIGN_Y         = p3["sign_y"]
    PART3_COMMA_Y        = p3["comma_y"]
    PART3_DIGIT_START_Y  = p3["digit_start_y"]
    PART3_DIGIT_STEP_Y   = p3["digit_step_y"]
    PART3_NUM_DIGIT_COLS = p3.get("num_digit_cols", 4)

    # SBD + Mã đề
    SBD_COLS_X       = t["sbd"]["cols_x"]
    MADE_COLS_X      = t["made"]["cols_x"]
    SBD_MADE_DIGIT_Y = t["sbd"]["digit_y"]

    # Rebuild derived data
    TEXT_ERASE_REGIONS = _build_text_erase_regions()
    ALL_BUBBLE_CENTERS = _collect_all_bubble_centers()
    ERASE_MASK         = _build_erase_mask()

    print(f"[OK] Template: {t.get('name', json_path)}")
    return t


def load_answers(json_path):
    """
    Đọc đáp án từ file JSON → trả về dict tương thích với process_sheet().
    Keys JSON là string ("1", "2"...) → chuyển thành int.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    correct = {}

    # Part I: {"1": "A", ...} → {1: "A", ...}
    if "part1" in raw:
        correct["part1"] = {int(k): v for k, v in raw["part1"].items()}

    # Part II: {"1": {"a": "Dung", ...}} → {1: {"a": "Dung", ...}}
    if "part2" in raw:
        correct["part2"] = {int(k): v for k, v in raw["part2"].items()}

    # Part III: {"1": "1234", ...} → {1: "1234", ...}
    if "part3" in raw:
        correct["part3"] = {int(k): v for k, v in raw["part3"].items()}

    print(f"[OK] Đáp án: {raw.get('exam_name', json_path)}")
    return correct


def save_result(result, output_dir="results", correct_answers=None):
    """
    Lưu kết quả chấm bài ra file JSON đầy đủ.
    result: dict trả về từ process_sheet()
    correct_answers: dict đáp án đúng (tùy chọn, để ghi kèm vào JSON)
    Trả về đường dẫn file JSON đã lưu.
    """
    if result is None:
        return None

    os.makedirs(output_dir, exist_ok=True)

    sbd = result.get("sbd", "unknown")
    made = result.get("made", "unknown")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    scores = result.get("scores", {})

    # Chuyển đáp án HS (key int → string cho JSON)
    out = {
        "timestamp": datetime.now().isoformat(),
        "sbd": sbd,
        "made": made,
        "score_total": result.get("score"),
        "max_score": result.get("max_score"),
        "score_part1": scores.get("part1"),
        "score_part2": scores.get("part2"),
        "score_part3": scores.get("part3"),
        "student_answers": {
            "part1": {str(k): v for k, v in result.get("part1", {}).items()},
            "part2": {str(k): v for k, v in result.get("part2", {}).items()},
            "part3": {str(k): v for k, v in result.get("part3", {}).items()},
        },
    }

    # Ghi kèm đáp án đúng nếu có
    if correct_answers:
        out["correct_answers"] = {
            "part1": {str(k): v for k, v in correct_answers.get("part1", {}).items()},
            "part2": {str(k): v for k, v in correct_answers.get("part2", {}).items()},
            "part3": {str(k): v for k, v in correct_answers.get("part3", {}).items()},
        }

    # Sanitize filename (SBD/MD có thể chứa '?' nếu không nhận được)
    safe_sbd = sbd.replace("?", "_")
    safe_made = made.replace("?", "_")
    fname = f"SBD_{safe_sbd}_MD_{safe_made}_{ts}.json"
    fpath = os.path.join(output_dir, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"[OK] JSON → {fpath}")
    return fpath


def export_excel(all_results, output_path="results/bang_diem.xlsx"):
    """
    Xuất bảng điểm tổng hợp ra file Excel.
    all_results: list of dict (mỗi phần tử là return value từ process_sheet)
                 hoặc dict {filename: result}
    output_path: đường dẫn file .xlsx
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("[LỖI] Cần cài openpyxl: pip install openpyxl")
        return None

    # Chuẩn hóa input
    if isinstance(all_results, dict):
        results_list = [v for v in all_results.values() if v is not None]
    else:
        results_list = [r for r in all_results if r is not None]

    if not results_list:
        print("[WARN] Không có kết quả để xuất Excel")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"

    # ── Styles ──
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    cell_align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ── Header row ──
    headers = [
        "STT", "SBD", "Mã đề", "Họ tên",
        "Phần I\n(/40)", "Phần II\n(/8)", "Phần III\n(/6)",
        "Tổng điểm\n(/54)", "Điểm 10",
    ]
    # Thêm cột đáp án Part I (Q1-Q40)
    for q in range(1, 41):
        headers.append(f"Q{q}")
    # Thêm cột Part III (Q1-Q6)
    for q in range(1, 7):
        headers.append(f"P3.Q{q}")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ──
    for ri, res in enumerate(sorted(results_list, key=lambda r: r.get("sbd", "")), start=1):
        scores = res.get("scores", {})
        s1 = scores.get("part1", 0) or 0
        s2 = scores.get("part2", 0) or 0
        s3 = scores.get("part3", 0) or 0
        total = res.get("score") or (s1 + s2 + s3)
        max_s = res.get("max_score") or 54
        diem10 = round(total / max_s * 10, 2) if max_s > 0 else 0

        row_data = [
            ri,
            res.get("sbd", ""),
            res.get("made", ""),
            "",  # Họ tên — user điền sau
            s1, s2, s3, total, diem10,
        ]
        # Đáp án Part I
        p1 = res.get("part1", {})
        for q in range(1, 41):
            row_data.append(p1.get(q, ""))
        # Đáp án Part III
        p3 = res.get("part3", {})
        for q in range(1, 7):
            row_data.append(p3.get(q, ""))

        row_num = ri + 1
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.border = thin_border
            if ci <= 4:
                cell.alignment = cell_align_left if ci == 4 else cell_align
            else:
                cell.alignment = cell_align

        # Highlight tổng điểm
        total_cell = ws.cell(row=row_num, column=8)
        if total >= 27:  # >= 50%
            total_cell.fill = green_fill
        else:
            total_cell.fill = red_fill

        # Highlight đáp án sai Part I (nếu tô trùng hoặc trống)
        for q in range(1, 41):
            ci = 9 + q  # column index (1-based)
            ans = p1.get(q, "")
            if ans in ("X", ""):
                ws.cell(row=row_num, column=ci).fill = red_fill

    # ── Column widths ──
    ws.column_dimensions["A"].width = 5    # STT
    ws.column_dimensions["B"].width = 12   # SBD
    ws.column_dimensions["C"].width = 8    # Mã đề
    ws.column_dimensions["D"].width = 25   # Họ tên
    ws.column_dimensions["E"].width = 9    # P1
    ws.column_dimensions["F"].width = 9    # P2
    ws.column_dimensions["G"].width = 9    # P3
    ws.column_dimensions["H"].width = 11   # Tổng
    ws.column_dimensions["I"].width = 9    # Điểm 10
    # Q1-Q40: narrow
    from openpyxl.utils import get_column_letter
    for ci in range(10, 10 + 40 + 6):
        ws.column_dimensions[get_column_letter(ci)].width = 5

    # Freeze header
    ws.freeze_panes = "A2"

    # ── Lưu ──
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"[OK] Excel → {output_path}")
    return output_path


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                    HÀM TIỆN ÍCH CƠ BẢN                             ║
# ╚════════════════════════════════════════════════════════════════════════╝

def order_points(pts):
    """Sắp xếp 4 điểm theo thứ tự: trên-trái, trên-phải, dưới-phải, dưới-trái."""
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]   # top-left: tổng nhỏ nhất
    rect[2] = pts[np.argmax(s)]   # bottom-right: tổng lớn nhất
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]  # top-right
    rect[3] = pts[np.argmax(diff)]  # bottom-left
    return rect


# ╔════════════════════════════════════════════════════════════════════════╗
# ║   BƯỚC 1-2: TỰ ĐỘNG PHÁT HIỆN GIẤY + NẮN THẲNG ẢNH               ║
# ║   auto_deskew_and_crop — Pipeline 2 lớp robust cho ảnh thực tế     ║
# ╚════════════════════════════════════════════════════════════════════════╝

# --- Tham số paper detection ---
_PAPER_MIN_AREA_RATIO = 0.10  # Giấy chiếm tối thiểu 10% ảnh (phone xa)
_PAPER_MAX_AREA_RATIO = 0.98  # Tối đa 98%
_PAPER_SIDE_RATIO_MIN = 0.35  # Cạnh đối diện chênh tối đa 65% (perspective)
_PAPER_ASPECT_MIN     = 0.45  # min(w,h)/max(w,h) — A4 dọc ≈ 0.73

# --- Tham số corner markers ---
_MARKER_MIN_AREA_RATIO = 0.0001  # Ô vuông nhỏ nhất (phone xa, ảnh lớn)
_MARKER_MAX_AREA_RATIO = 0.012   # Ô vuông lớn nhất

# --- Tham số refinement ---
_REFINE_MARKER_MIN = 0.00003     # Marker trên ảnh warped nhỏ hơn
_REFINE_MARKER_MAX = 0.004
_REFINE_MIN_SPAN   = 0.60        # 4 marker phải trải ≥60% ảnh warped
_REFINE_MARGIN     = 0.20        # Mỗi marker phải trong 20% từ góc ảnh


def _score_warp_quality(warped, method_name, corners):
    """
    Chấm điểm chất lượng ảnh warped — đa yếu tố thực tế:
      1) Sharpness (Laplacian variance)     — 40%  (quan trọng nhất)
      2) Paper coverage (vùng trắng)        — 25%
      3) Cleanliness (ít noise)             — 20%
      4) Corner stability (hình học)        — 10%
      5) Bonus nhẹ cho corner_markers       — +8
      6) Marker refinement thành công       — +15
    Returns (total_score, sharpness_raw, refined_corners_or_None)
    """
    gray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY) \
        if len(warped.shape) == 3 else warped.copy()
    h, w = gray.shape[:2]
    total_px = h * w
    score = 0.0
    detail = {}

    # ── 1) Sharpness — 40% (0-40) ──
    lap_var = cv2.Laplacian(gray, cv2.CV_64F).var()
    sharp_score = min(100.0, lap_var / 50.0) * 0.40
    score += sharp_score
    detail['sharp'] = f"{sharp_score:.1f}"

    # ── 2) Coverage — 25% (0-25) ──
    white_px = np.count_nonzero(gray > 150)
    coverage = (white_px / total_px) * 100.0 if total_px > 0 else 0
    cov_score = min(coverage, 100.0) * 0.25
    score += cov_score
    detail['cover'] = f"{cov_score:.1f}"

    # ── 3) Cleanliness — 20% (0-20) ──
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                   cv2.THRESH_BINARY_INV, 11, 2)
    noise_ratio = cv2.countNonZero(thresh) / total_px
    clean = max(0.0, 100.0 - noise_ratio * 80.0)
    clean_score = clean * 0.20
    score += clean_score
    detail['clean'] = f"{clean_score:.1f}"

    # ── 4) Corner stability — 10% (0-10) ──
    if corners is not None and len(corners) == 4:
        ordered = order_points(corners)
        tl, tr, br, bl = ordered
        w_top = np.linalg.norm(tr - tl)
        w_bot = np.linalg.norm(br - bl)
        h_left = np.linalg.norm(bl - tl)
        h_right = np.linalg.norm(br - tr)
        if w_top > 0 and w_bot > 0 and h_left > 0 and h_right > 0:
            w_sym = min(w_top, w_bot) / max(w_top, w_bot)
            h_sym = min(h_left, h_right) / max(h_left, h_right)
            symmetry = (w_sym + h_sym) / 2.0
            avg_w = (w_top + w_bot) / 2
            avg_h = (h_left + h_right) / 2
            aspect = min(avg_w, avg_h) / max(avg_w, avg_h)
            a4_fit = max(0.0, 1.0 - abs(aspect - 0.707) * 3.0)
            stab = (symmetry * 0.5 + a4_fit * 0.5) * 10.0
            score += stab
            detail['stab'] = f"{stab:.1f}"

    # ── 5) Bonus nhẹ cho corner_markers ──
    # [FIX] Đã loại bỏ +8 điểm tùy ý vì nó có thể làm corner_markers thắng 
    #       ngay cả khi paper_contour nắn chuẩn hơn.
    if method_name.startswith("corner_markers"):
        score += 0.0
        detail['bonus'] = '+0'

    # ── 6) Marker refinement ──
    refined = _refine_targeted(warped)
    if refined is not None:
        score += 15.0
        detail['refine'] = 'targeted+15'
    else:
        refined = _refine_with_markers(warped)
        if refined is not None:
            score += 8.0
            detail['refine'] = 'global+8'
        else:
            detail['refine'] = 'none'

    print(f"    score_detail: {detail} → {score:.1f}")
    return score, sharp_score, refined


def auto_deskew_and_crop(image, debug=False):
    """
    Tự động phát hiện phiếu trắc nghiệm và nắn thẳng.

    Luôn chạy CẢ 2 method, chọn kết quả tốt nhất:
      Method A : Tìm 4 ô vuông đen góc → warp → refine
      Method B : Tìm viền giấy (Canny / threshold / saturation)
                 → warp trung gian → refine bằng corner markers

    Returns:
      dict {
        "warped"      : ndarray  — ảnh nắn thẳng WARP_WIDTH × WARP_HEIGHT
        "corners"     : ndarray 4×2 float32 [TL, TR, BR, BL] trên ảnh gốc
        "method"      : str
        "success"     : bool — True nếu tìm được
        "debug_image" : ndarray | None
      }
    Raise ValueError nếu cả 2 đều thất bại.
    """
    debug_img = image.copy() if debug else None
    candidates = []  # (score, warped, corners, method_name)

    # ─── Method A: Corner markers ───
    markers = _find_corner_markers(image, debug_img)
    if markers is not None:
        ordered = order_points(markers)
        if _validate_marker_quad(ordered, image.shape[1], image.shape[0]):
            warped_a = _warp_to_rect(image, ordered)
            score_a, sharp_a, refined_a = _score_warp_quality(
                warped_a, "corner_markers", ordered)
            method_a = "corner_markers"
            corners_a = ordered
            if refined_a is not None:
                # ── FIX: single-warp thay vì double-warp ──
                refined_orig = _map_warped_to_original(refined_a, ordered)
                if _validate_marker_quad(refined_orig, image.shape[1], image.shape[0]):
                    warped_a_ref = _warp_to_rect(image, refined_orig)
                    score_a_ref, sharp_a_ref, _ = _score_warp_quality(
                        warped_a_ref, "corner_markers+refine", refined_orig)
                    # [FIX] Chỉ dùng refinement nếu score KHÔNG kém hơn
                    if score_a_ref >= score_a:
                        warped_a = warped_a_ref
                        score_a = score_a_ref
                        sharp_a = sharp_a_ref
                        method_a = "corner_markers+refine"
                        corners_a = refined_orig
            # [FIX] Bonus ĐIỀU KIỆN theo symmetry — tránh false positive.
            # Quad đối xứng cao (> 0.97) = detect đúng → +5 bonus
            # Thấp hơn → bonus giảm hoặc không có
            sym_a = _quad_symmetry(corners_a)
            if sym_a > 0.97:
                score_a += 5.0
                bonus_note = f"+5 direct, sym={sym_a:.3f}"
            elif sym_a > 0.94:
                score_a += 2.0
                bonus_note = f"+2 direct, sym={sym_a:.3f}"
            else:
                bonus_note = f"no bonus, sym={sym_a:.3f} (suspicious)"
            candidates.append((score_a, sharp_a, warped_a, corners_a, method_a))
            print(f"  [A] {method_a} → {score_a:.1f} ({bonus_note})")

    # ─── Method B: Paper contour ───
    paper = _find_paper_contour(image, debug_img)
    if paper is not None:
        ordered_p = order_points(paper)
        warped_b = _warp_to_rect(image, ordered_p)
        score_b, sharp_b, refined_b = _score_warp_quality(
            warped_b, "paper_contour", ordered_p)
        method_b = "paper_contour"
        corners_b = ordered_p
        if refined_b is not None:
            # ── FIX: single-warp thay vì double-warp ──
            refined_orig_b = _map_warped_to_original(refined_b, ordered_p)
            if _validate_marker_quad(refined_orig_b, image.shape[1], image.shape[0]):
                warped_b_ref = _warp_to_rect(image, refined_orig_b)
                score_b_ref, sharp_b_ref, _ = _score_warp_quality(
                    warped_b_ref, "paper+refine", refined_orig_b)
                # [FIX] Chỉ dùng refinement nếu score KHÔNG kém hơn
                if score_b_ref >= score_b:
                    warped_b = warped_b_ref
                    score_b = score_b_ref
                    sharp_b = sharp_b_ref
                    method_b = "paper+refine"
                    corners_b = refined_orig_b
        candidates.append((score_b, sharp_b, warped_b, corners_b, method_b))
        print(f"  [B] {method_b} → {score_b:.1f}")

    # ─── Method C: HYBRID paper + corner markers (DIRECT) ───
    # Cách 1 (tối ưu): Dùng paper corners làm HOA TIÊU, tìm marker 
    # TRỰC TIẾP trên ảnh gốc — KHÔNG cần warp trung gian.
    #
    # Quy trình:
    #   1) Paper contour cho 4 góc thô (mép giấy) trên ảnh gốc
    #   2) Crop vùng nhỏ BÊN TRONG giấy quanh mỗi góc thô
    #   3) Tìm ô vuông đen (marker) thực sự trong mỗi ROI
    #   4) Lấy TÂM marker → đó là 4 điểm warp chính xác pixel-level
    #   5) 1 WARP DUY NHẤT từ ảnh gốc đến tâm markers
    if paper is not None:
        ordered_paper = order_points(paper)
        img_h, img_w = image.shape[:2]
        gray_orig = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) \
            if len(image.shape) == 3 else image.copy()
        
        # Tâm giấy — dùng để xác định hướng "vào trong"
        paper_center_x = np.mean(ordered_paper[:, 0])
        paper_center_y = np.mean(ordered_paper[:, 1])
        
        # ROI rất nhỏ: marker nằm SÁT mép giấy (~2-5% kích thước giấy)
        paper_w = np.linalg.norm(ordered_paper[1] - ordered_paper[0])
        paper_h = np.linalg.norm(ordered_paper[3] - ordered_paper[0])
        # Vào trong: ~5% kích thước giấy (marker cách mép ~3%)
        inward_dist = max(30, int(min(paper_w, paper_h) * 0.055))
        # Ra ngoài: chỉ ~1.5% (marker có thể hơi lấn ra ngoài mép paper contour)
        outward_dist = max(10, int(min(paper_w, paper_h) * 0.015))
        
        marker_centers = []
        for i, corner_pt in enumerate(ordered_paper):
            cx_approx = int(corner_pt[0])
            cy_approx = int(corner_pt[1])
            
            # Hướng "vào trong" giấy
            dx_sign = int(np.sign(paper_center_x - cx_approx))
            dy_sign = int(np.sign(paper_center_y - cy_approx))
            
            # ROI: vào trong nhiều, ra ngoài ít
            x1 = max(0, cx_approx - (outward_dist if dx_sign > 0 else inward_dist))
            x2 = min(img_w, cx_approx + (inward_dist if dx_sign > 0 else outward_dist))
            y1 = max(0, cy_approx - (outward_dist if dy_sign > 0 else inward_dist))
            y2 = min(img_h, cy_approx + (inward_dist if dy_sign > 0 else outward_dist))
            
            roi = gray_orig[y1:y2, x1:x2]
            if roi.size == 0:
                break
            
            # Tìm ô vuông đen + ưu tiên vị trí GẦN góc paper nhất
            marker = _find_marker_near_corner(roi, cx_approx - x1, cy_approx - y1)
            if marker is None:
                break
            
            marker_x = x1 + marker[0]
            marker_y = y1 + marker[1]
            marker_centers.append([float(marker_x), float(marker_y)])
        
        if len(marker_centers) == 4:
            markers_orig = order_points(
                np.array(marker_centers, dtype="float32"))
            
            # Validate: markers phải tạo hình chữ nhật hợp lệ
            if _validate_marker_quad(markers_orig, img_w, img_h):
                # 1 WARP DUY NHẤT từ ảnh gốc dùng tâm markers
                warped_c = _warp_to_rect(image, markers_orig)
                
                # Score (KHÔNG refine thêm — đã chính xác pixel-level)
                score_c, sharp_c, _ = _score_warp_quality(
                    warped_c, "paper+markers", markers_orig)
                # [FIX] Bonus điều kiện theo symmetry
                sym_c = _quad_symmetry(markers_orig)
                if sym_c > 0.97:
                    score_c += 5.0
                    bonus_c = f"+5 cross-validate, sym={sym_c:.3f}"
                elif sym_c > 0.94:
                    score_c += 2.0
                    bonus_c = f"+2 cross-validate, sym={sym_c:.3f}"
                else:
                    bonus_c = f"no bonus, sym={sym_c:.3f} (suspicious)"
                candidates.append((score_c, sharp_c, warped_c, markers_orig,
                                   "paper+markers"))
                print(f"  [C] paper+markers → {score_c:.1f} ({bonus_c})")
            else:
                print(f"  [C] paper+markers: quad invalid → skip")
        else:
            print(f"  [C] paper+markers: only {len(marker_centers)}/4 markers → skip")

    # ─── Chọn kết quả tốt nhất ───
    if not candidates:
        raise ValueError("Không tìm được viền giấy lẫn 4 góc đen.")

    # Sort by score desc
    candidates.sort(key=lambda x: x[0], reverse=True)

    # [FIX] Tie-breaker dựa trên SYMMETRY thay vì sharpness.
    # Sharpness đã được tính 40% trong score rồi — không cần đếm 2 lần.
    # Symmetry cao hơn = quad chuẩn hơn = warp ít skew hơn.
    if len(candidates) > 1:
        top = candidates[0]
        runner = candidates[1]
        if abs(top[0] - runner[0]) < 2.0:
            sym_top = _quad_symmetry(top[3])
            sym_runner = _quad_symmetry(runner[3])
            if sym_runner > sym_top + 0.01:
                print(f"  [TIE-BREAK] scores within 2pts, "
                      f"picking {runner[4]} (more symmetric: "
                      f"{sym_runner:.3f} > {sym_top:.3f})")
                candidates[0], candidates[1] = candidates[1], candidates[0]

    best_score, best_sharp, best_warped, best_corners, best_method = candidates[0]

    if len(candidates) > 1:
        print(f"  [PICK] {best_method} ({best_score:.1f}) "
              f"over {candidates[1][4]} ({candidates[1][0]:.1f})")

    if debug_img is not None:
        _draw_debug(debug_img, best_corners, None, best_method.upper(),
                    (0, 165, 255))

    result = _result(best_warped, best_corners, best_method, True, debug_img)
    # Lưu danh sách tất cả candidates để process_sheet có thể thử lại
    result["_candidates"] = candidates
    return result


def _validate_marker_quad(ordered, img_w, img_h):
    """
    Kiểm tra 4 marker [TL, TR, BR, BL] có tạo thành hình chữ nhật hợp lệ:
    1) Convex
    2) Aspect ratio gần A4 (0.55 - 0.90)
    3) Cạnh đối diện không chênh quá 40%
    4) Chiếm >= 15% diện tích ảnh (tránh marker quá gần nhau)
    5) Góc trong hợp lý (60° - 120°)
    """
    tl, tr, br, bl = ordered
    # Cạnh
    w_top = np.linalg.norm(tr - tl)
    w_bot = np.linalg.norm(br - bl)
    h_left = np.linalg.norm(bl - tl)
    h_right = np.linalg.norm(br - tr)

    avg_w = (w_top + w_bot) / 2
    avg_h = (h_left + h_right) / 2
    if avg_w == 0 or avg_h == 0:
        return False

    # Aspect ratio: A4 dọc ≈ 0.707 (210/297mm)
    aspect = min(avg_w, avg_h) / max(avg_w, avg_h)
    if aspect < 0.62 or aspect > 0.85:
        return False

    # Cạnh đối diện không chênh quá 40%
    if w_top > 0 and w_bot > 0:
        if min(w_top, w_bot) / max(w_top, w_bot) < 0.60:
            return False
    if h_left > 0 and h_right > 0:
        if min(h_left, h_right) / max(h_left, h_right) < 0.60:
            return False

    # Diện tích quad >= 15% ảnh
    quad_area = cv2.contourArea(ordered)
    img_area = img_w * img_h
    if quad_area / img_area < 0.15:
        return False

    # Góc trong hợp lý (60° - 120°)
    for i in range(4):
        p1 = ordered[i]
        p2 = ordered[(i + 1) % 4]
        p3 = ordered[(i + 2) % 4]
        v1 = p1 - p2
        v2 = p3 - p2
        cos_a = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2) + 1e-8)
        angle = np.degrees(np.arccos(np.clip(cos_a, -1, 1)))
        if angle < 55 or angle > 125:
            return False

    return True


def _result(warped, corners, method, success, debug_img):
    """Helper tạo dict kết quả chuẩn."""
    return {
        "warped": warped,
        "corners": corners,
        "method": method,
        "success": success,
        "debug_image": debug_img,
    }


def _draw_debug(debug_img, paper_pts, marker_pts, label, color):
    """Vẽ thông tin debug lên ảnh gốc."""
    # Viền giấy (xanh lá)
    cv2.drawContours(debug_img, [paper_pts.astype(int)], -1, (0, 255, 0), 3)
    for pt in paper_pts:
        cx, cy = int(pt[0]), int(pt[1])
        box_half = 12
        cv2.rectangle(debug_img, (cx - box_half, cy - box_half),
                      (cx + box_half, cy + box_half), (0, 0, 255), 2)
    # Markers nếu có (cam)
    if marker_pts is not None:
        for pt in marker_pts:
            cv2.circle(debug_img, (int(pt[0]), int(pt[1])), 8, (0, 165, 255), -1)
    # Label
    y_lbl = max(15, int(paper_pts[0][1]) - 15)
    cv2.putText(debug_img, label, (int(paper_pts[0][0]), y_lbl),
                cv2.FONT_HERSHEY_SIMPLEX, 1.0, color, 2)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  LỚP 1: CORNER MARKERS — tìm 4 ô vuông đen ở 4 góc phiếu
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _find_corner_markers(image, debug_img=None,
                          min_ratio=_MARKER_MIN_AREA_RATIO,
                          max_ratio=_MARKER_MAX_AREA_RATIO):
    """
    Tìm 4 ô vuông đen ở 4 góc phiếu.
    Thử adaptive → simple threshold.
    Trả về 4 điểm float32 hoặc None.
    """
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    h, w = image.shape[:2]
    min_a = (w * h) * min_ratio
    max_a = (w * h) * max_ratio

    all_sq = []

    # Chiến lược 1: Adaptive threshold (chịu ánh sáng không đều)
    at = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                               cv2.THRESH_BINARY_INV, 31, 10)
    all_sq.extend(_extract_squares(at, min_a, max_a))

    # Chiến lược 2: Simple threshold (nhiều mức)
    if len(all_sq) < 4:
        for tval in [60, 80, 100, 120]:
            _, st = cv2.threshold(blurred, tval, 255, cv2.THRESH_BINARY_INV)
            for s in _extract_squares(st, min_a, max_a):
                if not any(abs(s[0]-e[0]) < 30 and abs(s[1]-e[1]) < 30
                           for e in all_sq):
                    all_sq.append(s)

    if len(all_sq) < 4:
        return None

    # Chọn 4 ứng viên gần 4 góc ảnh nhất
    all_sq.sort(key=lambda s: s[2], reverse=True)
    cands = np.array([(s[0], s[1]) for s in all_sq[:30]], dtype="float32")
    corners = _greedy_assign_corners(cands, w, h)

    if debug_img is not None:
        for pt in corners:
            cv2.circle(debug_img, (int(pt[0]), int(pt[1])), 10, (255, 0, 0), -1)
        cv2.putText(debug_img, "MARKERS", (int(corners[0][0]),
                    max(15, int(corners[0][1]) - 15)),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 0, 0), 2)
    return corners


def _extract_squares(thresh_img, min_a, max_a):
    """Tìm hình vuông đen trong ảnh binary. Trả về list (cx, cy, area)."""
    cnts, _ = cv2.findContours(thresh_img, cv2.RETR_EXTERNAL,
                               cv2.CHAIN_APPROX_SIMPLE)
    result = []
    for c in cnts:
        area = cv2.contourArea(c)
        if area < min_a or area > max_a:
            continue
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.04 * peri, True)
        if 4 <= len(approx) <= 6:
            x, y, bw, bh = cv2.boundingRect(approx)
            asp = bw / float(bh) if bh > 0 else 0
            if 0.45 < asp < 2.2:  # Rộng hơn cho perspective phone
                result.append((x + bw // 2, y + bh // 2, area))
    return result


def _greedy_assign_corners(candidates, img_w, img_h):
    """Gán mỗi góc ảnh cho ứng viên gần nhất (greedy, không trùng)."""
    targets = np.array([
        [0, 0], [img_w, 0], [img_w, img_h], [0, img_h]
    ], dtype="float32")
    chosen = []
    used = set()
    for t in targets:
        best_i, best_d = -1, float("inf")
        for i, c in enumerate(candidates):
            if i in used:
                continue
            d = np.linalg.norm(c - t)
            if d < best_d:
                best_d = d
                best_i = i
        chosen.append(candidates[best_i])
        used.add(best_i)
    return np.array(chosen, dtype="float32")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  LỚP 2: PAPER CONTOUR — tìm viền giấy trắng trên nền bất kỳ
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _find_paper_contour(image, debug_img=None):
    """
    Tìm viền giấy bằng nhiều chiến lược song song:
      A) Canny edge (nhiều mức low/high)
      B) Otsu threshold + morphology close
      C) Saturation channel (giấy trắng = saturation thấp)
    Trả về 4 góc giấy float32 hoặc None.
    """
    h, w = image.shape[:2]
    img_area = h * w
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Giảm kích thước nếu ảnh quá lớn (tăng tốc + giảm nhiễu)
    scale = 1.0
    if max(h, w) > 2000:
        scale = 2000.0 / max(h, w)
        small = cv2.resize(gray, None, fx=scale, fy=scale,
                           interpolation=cv2.INTER_AREA)
    else:
        small = gray.copy()

    blurred = cv2.GaussianBlur(small, (7, 7), 0)

    # Thu thập tất cả binary images từ nhiều chiến lược
    binaries = []

    # ── Chiến lược A: Canny edge (nhiều mức) ──
    for lo, hi in [(20, 60), (30, 100), (50, 150), (75, 200)]:
        edges = cv2.Canny(blurred, lo, hi)
        k = cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7))
        edges = cv2.dilate(edges, k, iterations=2)
        edges = cv2.erode(edges, k, iterations=1)
        binaries.append(edges)

    # ── Chiến lược B: Otsu threshold + morphology close ──
    _, otsu = cv2.threshold(blurred, 0, 255,
                            cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    k_close = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 15))
    closed = cv2.morphologyEx(otsu, cv2.MORPH_CLOSE, k_close, iterations=2)
    binaries.append(closed)

    # ── Chiến lược C: Saturation channel (giấy trắng = S thấp) ──
    hsv_small = cv2.cvtColor(
        cv2.resize(image, None, fx=scale, fy=scale,
                   interpolation=cv2.INTER_AREA)
        if scale < 1.0 else image,
        cv2.COLOR_BGR2HSV
    )
    sat = hsv_small[:, :, 1]
    _, sat_bin = cv2.threshold(sat, 40, 255, cv2.THRESH_BINARY_INV)
    k_sat = cv2.getStructuringElement(cv2.MORPH_RECT, (11, 11))
    sat_bin = cv2.morphologyEx(sat_bin, cv2.MORPH_CLOSE, k_sat, iterations=3)
    sat_bin = cv2.morphologyEx(sat_bin, cv2.MORPH_OPEN, k_sat, iterations=1)
    binaries.append(sat_bin)

    # ── Chiến lược D: Adaptive threshold block lớn ──
    ada = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                cv2.THRESH_BINARY, 51, 5)
    ada_inv = cv2.bitwise_not(ada)
    k_ada = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 15))
    ada_inv = cv2.morphologyEx(ada_inv, cv2.MORPH_CLOSE, k_ada, iterations=3)
    binaries.append(ada_inv)

    # Tìm quad tốt nhất từ tất cả binary images
    best_quad = None
    best_score = -1

    sh, sw = small.shape[:2]
    small_area = sh * sw

    for bimg in binaries:
        quad = _best_quad_from_binary(bimg, small_area)
        if quad is None:
            continue
        score = _score_paper_quad(quad, sw, sh, small_area)
        if score > best_score:
            best_quad = quad
            best_score = score

    if best_quad is None:
        return None

    # Scale ngược về kích thước gốc
    if scale < 1.0:
        best_quad = best_quad / scale

    return best_quad.astype("float32")


def _score_paper_quad(quad, img_w, img_h, img_area):
    """
    Tính điểm cho quad dựa trên:
    - Diện tích (lớn hơn = tốt, nhưng không phải lớn nhất)
    - Aspect ratio gần A4 (0.73) → bonus
    - Phạt nếu góc sát mép ảnh (< 1%) → có thể là cả ảnh, không phải giấy
    - Phạt nếu chiếm > 90% ảnh (quá lớn = bắt cả nền)
    """
    area = cv2.contourArea(quad)
    ratio = area / img_area

    # Base score = normalized area (0-1)
    base = ratio

    # Bonus: aspect ratio gần A4 (0.73)
    ordered = order_points(quad)
    tl, tr, br, bl = ordered
    avg_w = (np.linalg.norm(tr - tl) + np.linalg.norm(br - bl)) / 2
    avg_h = (np.linalg.norm(bl - tl) + np.linalg.norm(br - tr)) / 2
    if avg_w == 0 or avg_h == 0:
        return -1
    aspect = min(avg_w, avg_h) / max(avg_w, avg_h)
    a4_bonus = 1.0 - abs(aspect - 0.73) * 2.0  # Max 1.0 khi aspect=0.73
    a4_bonus = max(0.1, a4_bonus)

    # Phạt: góc sát mép ảnh (< 1% hoặc > 99%)
    edge_margin = 0.01
    edge_penalty = 1.0
    for pt in ordered:
        x_r = pt[0] / img_w if img_w > 0 else 0
        y_r = pt[1] / img_h if img_h > 0 else 0
        if x_r < edge_margin or x_r > (1 - edge_margin):
            edge_penalty *= 0.7
        if y_r < edge_margin or y_r > (1 - edge_margin):
            edge_penalty *= 0.7

    # Phạt quad quá lớn (> 90% ảnh → có thể là toàn bộ ảnh)
    size_penalty = 1.0
    if ratio > 0.90:
        size_penalty = 0.5
    elif ratio > 0.85:
        size_penalty = 0.8

    return base * a4_bonus * edge_penalty * size_penalty


def _best_quad_from_binary(binary, img_area):
    """
    Từ ảnh binary, tìm quadrilateral lớn nhất hợp lệ.
    Trả về quad 4×2 float32 hoặc None.
    """
    cnts, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL,
                               cv2.CHAIN_APPROX_SIMPLE)
    cnts = sorted(cnts, key=cv2.contourArea, reverse=True)

    for cnt in cnts[:15]:
        area = cv2.contourArea(cnt)
        ratio = area / img_area
        if ratio < _PAPER_MIN_AREA_RATIO or ratio > _PAPER_MAX_AREA_RATIO:
            continue

        peri = cv2.arcLength(cnt, True)
        # Thử nhiều epsilon xấp xỉ
        for eps in [0.015, 0.02, 0.03, 0.04, 0.06, 0.08]:
            approx = cv2.approxPolyDP(cnt, eps * peri, True)
            if len(approx) == 4:
                quad = approx.reshape(4, 2).astype("float32")
                if _is_valid_quad(quad, img_area):
                    return quad
                break  # Đã tìm được 4 cạnh, không thử eps khác

        # Nếu approxPolyDP không ra 4 điểm, thử convexHull + minAreaRect
        if len(approx) != 4:
            hull = cv2.convexHull(cnt)
            rect = cv2.minAreaRect(hull)
            box = cv2.boxPoints(rect).astype("float32")
            box_area = cv2.contourArea(box)
            if box_area / img_area >= _PAPER_MIN_AREA_RATIO:
                if _is_valid_quad(box, img_area):
                    return box

    return None


def _is_valid_quad(quad, img_area):
    """
    Kiểm tra tứ giác có hợp lệ làm viền giấy:
    - Diện tích hợp lý
    - Convex
    - Tỷ lệ cạnh hợp lý (phiếu A4)
    - Cạnh đối diện không chênh quá nhiều
    - Góc trong hợp lý (60°–120°)
    """
    area = cv2.contourArea(quad)
    ratio = area / img_area
    if ratio < _PAPER_MIN_AREA_RATIO or ratio > _PAPER_MAX_AREA_RATIO:
        return False

    if not cv2.isContourConvex(quad):
        return False

    ordered = order_points(quad)
    tl, tr, br, bl = ordered
    w_top = np.linalg.norm(tr - tl)
    w_bot = np.linalg.norm(br - bl)
    h_left = np.linalg.norm(bl - tl)
    h_right = np.linalg.norm(br - tr)

    avg_w = (w_top + w_bot) / 2
    avg_h = (h_left + h_right) / 2
    if avg_w == 0 or avg_h == 0:
        return False

    # Tỷ lệ khung
    aspect = min(avg_w, avg_h) / max(avg_w, avg_h)
    if aspect < _PAPER_ASPECT_MIN:
        return False

    # Cạnh đối diện không chênh quá nhiều
    if w_top > 0 and w_bot > 0:
        if min(w_top, w_bot) / max(w_top, w_bot) < _PAPER_SIDE_RATIO_MIN:
            return False
    if h_left > 0 and h_right > 0:
        if min(h_left, h_right) / max(h_left, h_right) < _PAPER_SIDE_RATIO_MIN:
            return False

    # Kiểm tra góc trong (tránh quad quá méo)
    for i in range(4):
        p1 = ordered[i]
        p2 = ordered[(i + 1) % 4]
        p3 = ordered[(i + 2) % 4]
        v1 = p1 - p2
        v2 = p3 - p2
        cos_a = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2) + 1e-8)
        angle = np.degrees(np.arccos(np.clip(cos_a, -1, 1)))
        if angle < 50 or angle > 140:
            return False

    return True


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  REFINEMENT — tìm corner markers trên ảnh warped trung gian
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _refine_with_markers(warped_raw):
    """Fallback: tìm markers bằng global contour. Giữ lại cho tương thích."""
    h, w = warped_raw.shape[:2]
    markers = _find_corner_markers(
        warped_raw, debug_img=None,
        min_ratio=_REFINE_MARKER_MIN, max_ratio=_REFINE_MARKER_MAX
    )
    if markers is None:
        return None
    ordered = order_points(markers)
    if not _validate_corner_positions(ordered, w, h):
        return None
    return ordered


def _refine_targeted(warped_raw):
    """
    Tìm 4 corner markers bằng TARGETED LOCAL SEARCH (SullyChen-inspired).

    Thay vì tìm global (nhiều false positive), chỉ tìm trong vùng góc.
    Kết hợp: contour detection + template matching.

    Trả về 4 góc ordered hoặc None.
    """
    gray = cv2.cvtColor(warped_raw, cv2.COLOR_BGR2GRAY) \
        if len(warped_raw.shape) == 3 else warped_raw.copy()
    h, w = gray.shape[:2]

    # Vùng tìm kiếm: 15% width, 8% height từ mỗi góc
    mx = int(w * 0.15)
    my = int(h * 0.08)

    corner_rois = [
        (0,      0,      mx, my),       # TL
        (w - mx, 0,      w,  my),       # TR
        (w - mx, h - my, w,  h),        # BR
        (0,      h - my, mx, h),        # BL
    ]

    found = []
    for (x1, y1, x2, y2) in corner_rois:
        roi = gray[y1:y2, x1:x2]
        marker = _find_marker_in_roi(roi)
        if marker is not None:
            found.append((x1 + marker[0], y1 + marker[1]))

    if len(found) != 4:
        return None

    ordered = order_points(np.array(found, dtype="float32"))
    if not _validate_corner_positions(ordered, w, h):
        return None

    return ordered


def _find_marker_near_corner(roi_gray, corner_x_in_roi, corner_y_in_roi):
    """
    Tìm ô vuông đen (corner marker) trong ROI, ƯU TIÊN vị trí GẦN góc paper.
    
    Giống _find_marker_in_roi nhưng thêm proximity scoring:
    - Marker xa góc paper > 40px bị phạt nặng
    - Ưu tiên ô vuông nhỏ, đặc (15-25px) gần mép giấy
    
    corner_x_in_roi, corner_y_in_roi: tọa độ góc paper trong hệ ROI
    Returns (cx, cy) hoặc None.
    """
    rh, rw = roi_gray.shape[:2]
    if rh < 15 or rw < 15:
        return None

    best_center = None
    best_score = 0

    for tval in [50, 70, 90, 110, 130]:
        _, binary = cv2.threshold(roi_gray, tval, 255, cv2.THRESH_BINARY_INV)
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL,
                                       cv2.CHAIN_APPROX_SIMPLE)
        for cnt in contours:
            area = cv2.contourArea(cnt)
            # Marker trên ảnh gốc: 60-2000px² (nhỏ hơn trên warped)
            if area < 40 or area > 3000:
                continue
            x, y, bw, bh = cv2.boundingRect(cnt)
            asp = bw / float(bh) if bh > 0 else 0
            if 0.45 < asp < 2.2:
                fill = area / (bw * bh) if bw * bh > 0 else 0
                if fill < 0.5:
                    continue
                squareness = 1.0 - abs(asp - 1.0) * 0.5
                
                # Tâm contour
                mcx = x + bw // 2
                mcy = y + bh // 2
                
                # Khoảng cách đến góc paper (trong hệ ROI)
                dist = np.sqrt((mcx - corner_x_in_roi)**2 + 
                               (mcy - corner_y_in_roi)**2)
                
                # [FIX] Proximity bonus — marker thật nằm CÁCH mép giấy 20-60px,
                # KHÔNG phải sát mép. Feature sát mép (dist<10px) thường là
                # shadow/nếp gấp/chữ in — phải bị phạt, không thưởng.
                #   dist < 10px  : nhiễu mép giấy → phạt mạnh (0.3)
                #   10-70px      : vùng marker hợp lý → bonus 1.0
                #   > 70px       : quá xa → giảm dần
                if dist < 10:
                    proximity = 0.3
                elif dist <= 70:
                    proximity = 1.0
                else:
                    proximity = max(0.2, 1.0 - (dist - 70) / 100.0)
                
                score = area * squareness * fill * proximity
                if score > best_score:
                    best_score = score
                    best_center = (mcx, mcy)

    if best_center is not None:
        return best_center

    # Fallback: Template matching (giữ nguyên)
    for tsize in [16, 20, 24]:
        pad = 5
        tmpl = np.ones((tsize + pad * 2, tsize + pad * 2), dtype=np.uint8) * 200
        tmpl[pad:pad + tsize, pad:pad + tsize] = 30

        if rh < tmpl.shape[0] or rw < tmpl.shape[1]:
            continue

        result = cv2.matchTemplate(roi_gray, tmpl, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, max_loc = cv2.minMaxLoc(result)

        if max_val > 0.35:
            return (max_loc[0] + tmpl.shape[1] // 2,
                    max_loc[1] + tmpl.shape[0] // 2)

    return None


def _find_marker_in_roi(roi_gray):
    """
    Tìm ô vuông đen (corner marker) trong ROI nhỏ.

    2 phương pháp:
      1) Contour detection (multi-threshold)
      2) Template matching (SullyChen-inspired) — fallback

    Returns (cx, cy) hoặc None.
    """
    rh, rw = roi_gray.shape[:2]
    if rh < 15 or rw < 15:
        return None

    # --- Method 1: Contour detection (nhiều ngưỡng) ---
    best_center = None
    best_score = 0

    for tval in [50, 70, 90, 110, 130]:
        _, binary = cv2.threshold(roi_gray, tval, 255, cv2.THRESH_BINARY_INV)
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL,
                                       cv2.CHAIN_APPROX_SIMPLE)
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < 60 or area > 4000:
                continue
            x, y, bw, bh = cv2.boundingRect(cnt)
            asp = bw / float(bh) if bh > 0 else 0
            if 0.45 < asp < 2.2:
                # Fill ratio: contourArea / boundingRectArea (vuông đặc ≈ 0.8+)
                fill = area / (bw * bh) if bw * bh > 0 else 0
                if fill < 0.5:
                    continue
                squareness = 1.0 - abs(asp - 1.0) * 0.5
                score = area * squareness * fill
                if score > best_score:
                    best_score = score
                    best_center = (x + bw // 2, y + bh // 2)

    if best_center is not None:
        return best_center

    # --- Method 2: Template matching (SullyChen-inspired) ---
    for tsize in [16, 20, 24]:
        pad = 5
        tmpl = np.ones((tsize + pad * 2, tsize + pad * 2), dtype=np.uint8) * 200
        tmpl[pad:pad + tsize, pad:pad + tsize] = 30

        if rh < tmpl.shape[0] or rw < tmpl.shape[1]:
            continue

        result = cv2.matchTemplate(roi_gray, tmpl, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, max_loc = cv2.minMaxLoc(result)

        if max_val > 0.35:
            return (max_loc[0] + tmpl.shape[1] // 2,
                    max_loc[1] + tmpl.shape[0] // 2)

    return None


def _validate_corner_positions(ordered_pts, img_w, img_h):
    """
    Kiểm tra 4 điểm [TL, TR, BR, BL] có ở gần 4 góc ảnh:
    1) Span tổng ≥ _REFINE_MIN_SPAN
    2) Mỗi điểm trong _REFINE_MARGIN % từ góc tương ứng
    """
    tl, tr, br, bl = ordered_pts
    xs = [tl[0], tr[0], br[0], bl[0]]
    ys = [tl[1], tr[1], br[1], bl[1]]

    span_w = (max(xs) - min(xs)) / img_w if img_w > 0 else 0
    span_h = (max(ys) - min(ys)) / img_h if img_h > 0 else 0
    if span_w < _REFINE_MIN_SPAN or span_h < _REFINE_MIN_SPAN:
        return False

    mx = img_w * _REFINE_MARGIN
    my = img_h * _REFINE_MARGIN
    expected = [(0, 0), (img_w, 0), (img_w, img_h), (0, img_h)]
    for pt, (ex, ey) in zip([tl, tr, br, bl], expected):
        if abs(pt[0] - ex) > mx or abs(pt[1] - ey) > my:
            return False

    return True


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  WARP HELPER + COMPATIBILITY WRAPPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _warp_to_rect(image, corners):
    """Nắn ảnh nghiêng → hình chữ nhật WARP_WIDTH × WARP_HEIGHT."""
    dst = np.array([
        [0, 0], [WARP_WIDTH - 1, 0],
        [WARP_WIDTH - 1, WARP_HEIGHT - 1], [0, WARP_HEIGHT - 1]
    ], dtype="float32")
    M = cv2.getPerspectiveTransform(corners, dst)
    return cv2.warpPerspective(image, M, (WARP_WIDTH, WARP_HEIGHT))


def _quad_symmetry(corners):
    """
    Đo độ đối xứng của quad 4 góc [TL, TR, BR, BL] (ordered).
    Returns giá trị 0-1, với 1 = hoàn hảo (hình chữ nhật đều).
    Dùng để phát hiện false positive (quad bất đối xứng → detect sai).
    """
    tl, tr, br, bl = corners
    w_top = np.linalg.norm(tr - tl)
    w_bot = np.linalg.norm(br - bl)
    h_left = np.linalg.norm(bl - tl)
    h_right = np.linalg.norm(br - tr)
    if min(w_top, w_bot, h_left, h_right) <= 0:
        return 0.0
    w_sym = min(w_top, w_bot) / max(w_top, w_bot)
    h_sym = min(h_left, h_right) / max(h_left, h_right)
    return float((w_sym + h_sym) / 2.0)


def _map_warped_to_original(refined_pts, original_corners):
    """
    Map tọa độ từ warped space (1400×1920) ngược về ảnh gốc.

    Dùng nghịch đảo ma trận perspective để tránh double-warp.
    refined_pts:      4 điểm ordered trong warped space
    original_corners: 4 góc ordered dùng cho warp ban đầu
    Returns: 4 điểm tương ứng trên ảnh gốc (float32, ordered)
    """
    dst = np.array([
        [0, 0], [WARP_WIDTH - 1, 0],
        [WARP_WIDTH - 1, WARP_HEIGHT - 1], [0, WARP_HEIGHT - 1]
    ], dtype="float32")
    M = cv2.getPerspectiveTransform(original_corners, dst)
    M_inv = np.linalg.inv(M)
    pts = refined_pts.reshape(-1, 1, 2).astype("float64")
    mapped = cv2.perspectiveTransform(pts, M_inv)
    return order_points(mapped.reshape(-1, 2).astype("float32"))


# Alias cho code cũ — detect_paper_and_warp gọi auto_deskew_and_crop
def detect_paper_and_warp(image, debug=False):
    """Alias tương thích: gọi auto_deskew_and_crop."""
    return auto_deskew_and_crop(image, debug=debug)


def detect_corners(image):
    """Wrapper tương thích: trả về 4 góc [TL, TR, BR, BL]."""
    result = auto_deskew_and_crop(image, debug=False)
    return result["corners"]


def warp_perspective(image, corners):
    """Wrapper tương thích: nắn ảnh từ 4 góc đã cho."""
    return _warp_to_rect(image, corners)


# ╔════════════════════════════════════════════════════════════════════════╗
# ║   LỚP 1: XÓA CHỮ IN TRÊN ẢNH WARPED (Smart Erase + Punch-holes)  ║
# ╚════════════════════════════════════════════════════════════════════════╝

def erase_printed_text(warped_img):
    """
    LỚP 1 (MẠNH NHẤT): Tô TRẮNG chữ in trên ảnh warped TRƯỚC threshold.

    Kỹ thuật punch-holes:
      - Phủ HCN trắng lên TOÀN BỘ vùng text (kể cả chồng lên bubble)
      - Nhưng ĐỤC LỖ TRÒN (punch-holes) tại mỗi bubble center
      - → Text GIỮA các bubble bị xóa triệt để
      - → Bubble được BẢO VỆ nguyên vẹn

    Ưu điểm: có thể phủ vùng rất rộng (y=590-680 cho header ABCD)
    mà KHÔNG mất dữ liệu bubble row 1.
    """
    result = warped_img.copy()
    # Áp dụng pre-computed mask: nơi ERASE_MASK=255 → tô trắng
    result[ERASE_MASK == 255] = [255, 255, 255]
    return result


# ╔════════════════════════════════════════════════════════════════════════╗
# ║   BƯỚC 3: TIỀN XỬ LÝ — 3 LỚP BẢO VỆ                              ║
# ╚════════════════════════════════════════════════════════════════════════╝

def preprocess(warped, enhance_camera=False):
    """
    Tiền xử lý: trả về ảnh xám (blur) cho detection.

    Nâng cấp: Illumination Flattening (Làm phẳng ánh sáng)
    - Giúp triệt tiêu bóng đổ và chênh lệch sáng tối giữa các vùng trên phiếu.
    - Giúp bút chì nhạt trở nên nổi bật hơn sau khi chia cho nền.
    """
    # --- LỚP 1: erase_printed_text (punch-hole) ---
    warped_clean = erase_printed_text(warped)

    gray_raw = cv2.cvtColor(warped_clean, cv2.COLOR_BGR2GRAY)

    # --- Làm phẳng nền giấy (Illumination Flattening / Background Division) ---
    # Dùng MORPH_CLOSE để ước lượng 'độ sáng của nền' (bỏ qua các vết mực/bút chì)
    bg_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (51, 51))
    background = cv2.morphologyEx(gray_raw, cv2.MORPH_CLOSE, bg_kernel)
    
    # Chia ảnh gốc cho nền để triệt tiêu chênh lệch ánh sáng cục bộ
    # flat_gray = (gray_raw / background) * 255
    flat_gray = cv2.divide(gray_raw, background, scale=255)

    # Tăng cường độ tương phản (Histogram Equalization nhẹ) nếu cần
    if enhance_camera:
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        flat_gray = clahe.apply(flat_gray)

    gray = cv2.GaussianBlur(flat_gray, (5, 5), 0)

    # Binary CHỈ cho debug visualization
    thresh = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 15, 8
    )
    kernel = cv2.getStructuringElement(
        cv2.MORPH_ELLIPSE, (MORPH_KERNEL_SIZE, MORPH_KERNEL_SIZE)
    )
    cleaned = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)

    return gray, thresh, cleaned


def detect_section_offsets(gray):
    """
    Detect ô vuông đen (■) trên ảnh warped để tính offset cục bộ.

    Mỗi section (Part I, II, III) có các ô vuông đen ở viền trên.
    So sánh vị trí thực tế vs mong đợi → tính (dy) cho từng section.
    Giúp xử lý ảnh bị phồng (paper bulge) mà 4-corner warp không fix được.

    Returns: dict {"part1": dy1, "part2": dy2, "part3": dy3}
    """
    h, w = gray.shape[:2]

    # Vị trí MONG ĐỢI của markers (y, x) — ô vuông đen ở viền trên mỗi section
    # Đo từ 2 ảnh mẫu warp chuẩn (te2.jpg + 30-04-06-TL)
    EXPECTED_MARKERS = {
        "part1": {"y": 600,  "xs": [350, 700]},
        "part2": {"y": 1054, "xs": [350, 700, 1050]},
        "part3": {"y": 1340, "xs": [233, 466, 700, 933, 1166]},
    }

    SEARCH_H = 60      # tìm kiếm ±60px theo y (phone bị lệch nhiều hơn)
    SEARCH_W = 35      # tìm kiếm ±35px theo x

    offsets = {"part1": 0, "part2": 0, "part3": 0}

    for section, info in EXPECTED_MARKERS.items():
        exp_y = info["y"]
        xs = info["xs"]
        detected_ys = []

        for exp_x in xs:
            # Crop vùng tìm kiếm
            y1 = max(0, exp_y - SEARCH_H)
            y2 = min(h, exp_y + SEARCH_H)
            x1 = max(0, exp_x - SEARCH_W)
            x2 = min(w, exp_x + SEARCH_W)
            roi = gray[y1:y2, x1:x2]
            if roi.size == 0:
                continue

            # Tìm ô vuông đen: threshold thấp (< 80 = rất tối)
            _, binary = cv2.threshold(roi, 80, 255, cv2.THRESH_BINARY_INV)
            contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)

            best_cy = None
            best_score = 0
            for cnt in contours:
                area = cv2.contourArea(cnt)
                # Ô vuông ~20x19 = ~350px², chấp nhận 80-500
                if 80 < area < 500:
                    x_c, y_c, cw, ch = cv2.boundingRect(cnt)
                    # Phải gần vuông
                    aspect = max(cw, ch) / max(min(cw, ch), 1)
                    if aspect < 2.0:
                        cy_abs = y1 + y_c + ch // 2
                        if area > best_score:
                            best_score = area
                            best_cy = cy_abs

            if best_cy is not None:
                detected_ys.append(best_cy)

        if detected_ys:
            avg_y = sum(detected_ys) / len(detected_ys)
            dy = avg_y - exp_y
            offsets[section] = round(dy)

    return offsets


def _cluster_1d(values, max_gap):
    """Cluster 1D values by max gap, return cluster centers."""
    if not values:
        return []
    vals = sorted(values)
    clusters = [[vals[0]]]
    for v in vals[1:]:
        if v - clusters[-1][-1] <= max_gap:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    return [int(round(sum(c) / len(c))) for c in clusters]


def detect_part3_offset_from_digits(gray):
    """Estimate Part III vertical offset from digit bubble rows."""
    h, w = gray.shape[:2]

    cols = []
    for blk in PART3_BLOCKS:
        cols.append(blk["sign_x"])
        cols.extend(blk["cols_x"])

    if not cols:
        return None

    x_min = max(0, int(min(cols) - BUBBLE_RADIUS * 3))
    x_max = min(w, int(max(cols) + BUBBLE_RADIUS * 3))
    y_min = max(0, int(PART3_SIGN_Y - BUBBLE_RADIUS * 3))
    y_max = min(h, int(PART3_DIGIT_START_Y + 9 * PART3_DIGIT_STEP_Y + BUBBLE_RADIUS * 3))

    roi = gray[y_min:y_max, x_min:x_max]
    if roi.size == 0:
        return None

    roi_blur = cv2.medianBlur(roi, 5)
    circles = cv2.HoughCircles(
        roi_blur,
        cv2.HOUGH_GRADIENT,
        dp=1,
        minDist=20,
        param1=50,
        param2=25,
        minRadius=8,
        maxRadius=18,
    )
    if circles is None:
        return None

    ys = [int(c[1]) + y_min for c in circles[0]]
    row_centers = _cluster_1d(ys, max_gap=10)
    if not row_centers:
        return None

    digit_rows = [y for y in row_centers if y >= PART3_DIGIT_START_Y - 20]
    if not digit_rows:
        digit_rows = row_centers

    expected_rows = [int(PART3_DIGIT_START_Y + d * PART3_DIGIT_STEP_Y) for d in range(10)]
    offsets = []
    for exp in expected_rows:
        nearest = min(digit_rows, key=lambda y: abs(y - exp))
        if abs(nearest - exp) <= 25:
            offsets.append(nearest - exp)

    if len(offsets) < 4:
        return None

    offset = int(round(float(np.median(offsets))))
    # Clamp to avoid outlier shift from bad circle detection.
    return max(-12, min(12, offset))


def mask_printed_text(thresh_img):
    """
    LỚP 3 (safety net): Tô ĐEN vùng text trên ảnh threshold.
    Cũng dùng ERASE_MASK với punch-holes để bảo vệ bubble.
    """
    result = thresh_img.copy()
    # Nơi ERASE_MASK=255 → tô đen (=0) trên threshold
    result[ERASE_MASK == 255] = 0
    return result


# ╔════════════════════════════════════════════════════════════════════════╗
# ║        BƯỚC 4: KIỂM TRA BUBBLE CÓ ĐƯỢC TÔ ĐEN KHÔNG               ║
# ╚════════════════════════════════════════════════════════════════════════╝

def is_bubble_filled(gray_img, cx, cy, radius=BUBBLE_RADIUS,
                     threshold=FILL_THRESHOLD, check_circularity=True):
    """
    Kiểm tra bubble tại (cx, cy) bằng hệ tọa độ NGƯỠNG ĐỘNG (Adaptive Threshold).
    
    Cơ chế: 
    1. Lấy vùng ảnh bao quanh bubble (ROI).
    2. Tính độ sáng trung bình của 'vòng nhẫn' (ring) bao quanh bubble để tìm mức TRẮNG CỤC BỘ.
    3. Tính độ sáng trung bình của lõi Bubble.
    4. Fill Ratio = 1 - (Bubble_Mean / Local_White_Mean).
    
    Ưu điểm: Nếu 1 góc phiếu bị tối (shadow), Local_White_Mean sẽ giảm xuống, 
    giúp phép chia vẫn cho ra kết quả Ratio chuẩn xác, không bị chấm nhầm.
    """
    h, w = gray_img.shape[:2]
    # Mở rộng vùng lấy mẫu để lấy được cả vùng trắng xung quanh bubble
    pad = 12 
    x1 = max(0, int(cx - radius - pad))
    y1 = max(0, int(cy - radius - pad))
    x2 = min(w, int(cx + radius + pad))
    y2 = min(h, int(cy + radius + pad))

    roi = gray_img[y1:y2, x1:x2]
    if roi.size == 0 or roi.shape[0] < radius*2 or roi.shape[1] < radius*2:
        return False, 0.0

    # Tọa độ tương đối của tâm trong ROI
    rx, ry = int(cx - x1), int(cy - y1)

    # 1. Tạo mask cho lõi BUBBLE (để tính độ đen)
    bubble_mask = np.zeros(roi.shape[:2], dtype=np.uint8)
    cv2.circle(bubble_mask, (rx, ry), radius, 255, -1)
    bubble_pixels = roi[bubble_mask == 255]
    if len(bubble_pixels) == 0: return False, 0.0
    bubble_mean = np.mean(bubble_pixels)

    # 2. Tạo mask cho VÒNG NHẪN (để tính độ trắng của giấy xung quanh)
    # Lấy từ radius+4 đến radius+10
    ring_mask = np.zeros(roi.shape[:2], dtype=np.uint8)
    cv2.circle(ring_mask, (rx, ry), radius + 10, 255, -1)
    cv2.circle(ring_mask, (rx, ry), radius + 4, 0, -1)
    ring_pixels = roi[ring_mask == 255]
    
    # Nếu không lấy được mẫu lề (sát mép ảnh), dùng chuẩn 255 hoặc trung bình ảnh
    local_white = np.mean(ring_pixels) if ring_pixels.size > 10 else 255.0
    # Đảm bảo local_white không quá thấp gây lỗi chia cho 0 hoặc noise
    local_white = max(local_white, 50.0)

    # 3. Tính tỉ lệ lấp đầy thích ứng
    # Ratio càng cao -> càng đen hơn so với giấy xung quanh
    ratio = 1.0 - (bubble_mean / local_white)
    
    # Clip ratio trong khoảng [0, 1]
    ratio = max(0.0, min(1.0, ratio))

    if ratio <= threshold:
        return False, ratio

    return True, ratio


def _crop_bubble_for_cnn(gray_img, cx, cy, radius=BUBBLE_RADIUS, crop_size=CNN_IMG_SIZE):
    """Crop a square around bubble center for CNN input."""
    h, w = gray_img.shape[:2]
    pad = radius + 4
    x1 = max(0, int(cx - pad))
    y1 = max(0, int(cy - pad))
    x2 = min(w, int(cx + pad))
    y2 = min(h, int(cy + pad))
    roi = gray_img[y1:y2, x1:x2]
    if roi.size == 0:
        return None
    return cv2.resize(roi, (crop_size, crop_size), interpolation=cv2.INTER_AREA)


def _load_bubble_cnn():
    """Lazy-load CNN model if available."""
    global _CNN_MODEL, _CNN_DEVICE, _CNN_READY, _CNN_ERROR
    if _CNN_READY or _CNN_ERROR is not None:
        return
    if not HYBRID_CNN_ENABLE:
        _CNN_ERROR = "disabled"
        return
    if not os.path.exists(BUBBLE_CNN_PATH):
        _CNN_ERROR = "model_not_found"
        return
    try:
        import torch
        from grading.engine.train_bubble_cnn import BubbleCNN
    except Exception as exc:
        _CNN_ERROR = str(exc)
        return

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = BubbleCNN().to(device)
    try:
        state = torch.load(BUBBLE_CNN_PATH, map_location=device, weights_only=True)
    except TypeError:
        state = torch.load(BUBBLE_CNN_PATH, map_location=device)
    model.load_state_dict(state)
    model.eval()

    _CNN_MODEL = model
    _CNN_DEVICE = device
    _CNN_READY = True


def _predict_bubble_cnn(gray_img, cx, cy):
    """Predict fill probability using CNN. Returns None if unavailable."""
    _load_bubble_cnn()
    if not _CNN_READY:
        return None

    crop = _crop_bubble_for_cnn(gray_img, cx, cy)
    if crop is None:
        return None

    try:
        import torch
        img = crop.astype(np.float32) / 255.0
        # Ensure 2D grayscale input → (1, 1, H, W)
        if img.ndim == 3:
            img = img[:, :, 0]  # drop channels if accidentally color
        tensor = torch.from_numpy(img).unsqueeze(0).unsqueeze(0).to(_CNN_DEVICE)

        with torch.no_grad():
            out = _CNN_MODEL(tensor)
            probs = torch.softmax(out, dim=1)
            return float(probs[0, 1].item())
    except Exception as e:
        print(f"[CNN WARN] _predict_bubble_cnn failed: {e}")
        return None


HYBRID_ALWAYS_CNN = False   # Only run CNN on ambiguous bubbles (saves ~80% time)

def _hybrid_score(gray_img, cx, cy, threshold=FILL_THRESHOLD, force_cnn=False):
    """Return (score, ratio, cnn_conf) using OpenCV + CNN.
    When HYBRID_ALWAYS_CNN=True, CNN runs on ALL bubbles and final score = max(ratio, cnn).
    """
    _, ratio = is_bubble_filled(gray_img, cx, cy, threshold=threshold)
    cnn_conf = None

    use_cnn = HYBRID_CNN_ENABLE and (
        HYBRID_ALWAYS_CNN or force_cnn or (HYBRID_RATIO_LOW <= ratio <= HYBRID_RATIO_HIGH)
    )
    if use_cnn:
        cnn_conf = _predict_bubble_cnn(gray_img, cx, cy)

    if cnn_conf is None:
        return ratio, ratio, None

    score = max(ratio, cnn_conf)
    return score, ratio, cnn_conf


def _find_p3_box_near(gray_img, cx, cy, box_size=P3_OCR_BOX_SIZE):
    """Find a handwriting box near (cx, cy). Returns (x1,y1,x2,y2) or None."""
    h, w = gray_img.shape[:2]
    half = int(box_size / 2)
    search = int(box_size * 2.0)
    x1 = max(0, int(cx - search))
    y1 = max(0, int(cy - search))
    x2 = min(w, int(cx + search))
    y2 = min(h, int(cy + search))
    roi = gray_img[y1:y2, x1:x2]
    if roi.size == 0:
        return None

    thr = cv2.adaptiveThreshold(
        roi, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 31, 7
    )
    thr = cv2.medianBlur(thr, 3)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    thr = cv2.dilate(thr, kernel, iterations=1)
    cnts, _ = cv2.findContours(thr, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    best = None
    for c in cnts:
        area = cv2.contourArea(c)
        if area < 120 or area > 5000:
            continue
        x, y, bw, bh = cv2.boundingRect(c)
        aspect = bw / float(bh) if bh > 0 else 0
        if 0.5 <= aspect <= 1.8:
            cx_c = x + bw / 2
            cy_c = y + bh / 2
            dist = (cx_c - (cx - x1)) ** 2 + (cy_c - (cy - y1)) ** 2
            if best is None or dist < best[0]:
                best = (dist, x, y, bw, bh)

    if best is None:
        return None

    _, bx, by, bw, bh = best
    gx1 = max(0, int(x1 + bx))
    gy1 = max(0, int(y1 + by))
    gx2 = min(w, int(x1 + bx + bw))
    gy2 = min(h, int(y1 + by + bh))
    return gx1, gy1, gx2, gy2


def _default_p3_box(gray_img, cx, cy, box_size=P3_OCR_BOX_SIZE):
    """Fallback box centered at (cx, cy)."""
    h, w = gray_img.shape[:2]
    half = int(box_size / 2)
    x1 = max(0, int(cx - half))
    y1 = max(0, int(cy - half))
    x2 = min(w, int(cx + half))
    y2 = min(h, int(cy + half))
    return x1, y1, x2, y2


def _ocr_digit_from_box(gray_img, box):
    """OCR a single digit from a handwriting box. Returns (digit, ink_ratio)."""
    if box is None:
        return -1, 0.0
    try:
        import pytesseract
    except Exception:
        return -1, 0.0

    x1, y1, x2, y2 = box
    roi = gray_img[y1:y2, x1:x2]
    if roi.size == 0:
        return -1, 0.0

    # Remove box border by cropping inner region.
    h, w = roi.shape[:2]
    margin = max(2, int(min(h, w) * 0.1))
    roi = roi[margin:h - margin, margin:w - margin]
    if roi.size == 0:
        return -1, 0.0

    # Resize to help OCR.
    roi = cv2.resize(roi, (0, 0), fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC)
    blur = cv2.GaussianBlur(roi, (3, 3), 0)
    _, thr = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    ink_ratio = float(np.count_nonzero(thr)) / float(thr.size)
    if ink_ratio < P3_OCR_INK_MIN or ink_ratio > P3_OCR_INK_MAX:
        return -1, ink_ratio
    thr = cv2.medianBlur(thr, 3)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    thr = cv2.dilate(thr, kernel, iterations=1)
    config = "--psm 10 -c tessedit_char_whitelist=0123456789"
    txt = pytesseract.image_to_string(thr, config=config)
    if not txt:
        txt = pytesseract.image_to_string(roi, config=config)
    txt = (txt or "").strip()
    if len(txt) != 1:
        return -1, ink_ratio
    ch = txt[0]
    return (int(ch) if ch.isdigit() else -1), ink_ratio


# ╔════════════════════════════════════════════════════════════════════════╗
# ║      BƯỚC 5a: TRÍCH XUẤT ĐÁP ÁN PHẦN I (40 câu ABCD)             ║
# ╚════════════════════════════════════════════════════════════════════════╝

def _detect_filled_choices(ratios):
    """
    Phát hiện bubble được tô — 2 pha:

    Pha 1 (Absolute): raw > FILL_THRESHOLD (0.38)
      → Dùng cho ảnh sạch, scan. Có thể detect nhiều → X.
      → FILL_THRESHOLD cao (0.38) loại bỏ baseline inflate (~0.34).

    Pha 2 (Adaptive): CHỈ chạy khi Pha 1 không tìm thấy gì.
      → Trừ noise_floor (2nd smallest) → tìm bubble nổi trội nhất.
      → Điều kiện: top_adj > 0.05 VÀ (top_adj - 2nd_adj) > 0.03
      → Không bao giờ return nhiều → không gây false X.

    Tự thích ứng: scan sạch → Pha 1 xử lý. Phone nhạt → Pha 2 bắt.
    """
    if not ratios:
        return []

    vals = list(ratios.values())
    if max(vals) < 0.12:
        return []

    # --- Pha 1: Absolute ---
    filled = [ch for ch, r in ratios.items() if r > FILL_THRESHOLD]
    if filled:
        if len(filled) == 1:
            return filled
        # Pha 1b: Nhiều bubble vượt threshold → tìm dominant
        # Trường hợp: 1 chấm nhỏ (~0.40) + 1 tô đậm (~0.60+)
        # → chọn cái đậm nhất nếu gap đủ lớn
        filled_ratios = sorted([(ch, ratios[ch]) for ch in filled],
                               key=lambda x: x[1], reverse=True)
        top_ch, top_r = filled_ratios[0]
        second_r = filled_ratios[1][1]
        gap = top_r - second_r
        # Dominant nếu: gap > 0.12 HOẶC top gấp >1.4x second
        if gap > 0.12 or (second_r > 0 and top_r / second_r > 1.4):
            return [top_ch]  # Chọn đáp án đậm nhất
        return filled  # Thật sự tô nhiều → X

    # --- Pha 2: Adaptive (chỉ khi Pha 1 trống) ---
    sorted_vals = sorted(vals)
    noise_floor = sorted_vals[1] if len(sorted_vals) >= 3 else sorted_vals[0]

    adjusted = [(ch, max(0.0, r - noise_floor)) for ch, r in ratios.items()]
    adjusted.sort(key=lambda x: x[1], reverse=True)

    top_ch, top_adj = adjusted[0]
    second_adj = adjusted[1][1] if len(adjusted) > 1 else 0.0

    if top_adj > 0.05 and (top_adj - second_adj) > 0.03:
        return [top_ch]

    return []


def _pick_answer(filled_choices):
    """Convert filled_choices list → answer string."""
    if len(filled_choices) == 1:
        return filled_choices[0]
    elif len(filled_choices) > 1:
        return "X"
    return ""


def _count_detected(answers_dict):
    """Count how many questions have a real answer (not '' or 'X')."""
    return sum(1 for a in answers_dict.values()
               if isinstance(a, str) and a not in ("", "X")
               or isinstance(a, dict) and any(v not in ("", "X") for v in a.values()))


def extract_part1(cleaned_img, y_offset=0):
    """
    Đọc 40 câu trắc nghiệm ABCD.
    y_offset: bù lệch y do ảnh phồng (từ detect_section_offsets).
    Hybrid 1-pass: max(OpenCV, CNN) cho mỗi bubble.
    Trả về:
      answers: {1: 'A', 2: 'C', ...}  ('X'=tô nhiều, ''=không tô)
      details: {1: {'A': 0.05, 'B': 0.72, ...}, ...}  (fill ratio)
    """
    answers = {}
    details = {}

    for cfg in PART1_COLS:
        sx, sy = cfg["start_x"], cfg["start_y"]
        dx, dy = cfg["step_x"], cfg["step_y"]
        q_start = cfg["q_start"]

        col_rows = cfg.get("num_rows", PART1_NUM_ROWS)
        for row in range(col_rows):
            q = q_start + row
            cy = sy + row * dy + y_offset
            ratios = {}

            for ci, choice in enumerate(PART1_CHOICES):
                cx = sx + ci * dx
                score, ratio, cnn_conf = _hybrid_score(cleaned_img, cx, cy)
                ratios[choice] = round(score, 3)

            details[q] = ratios
            filled_choices = _detect_filled_choices(ratios)
            ans = _pick_answer(filled_choices)
            answers[q] = ans

            # Debug: log fill ratios cho câu bị blank hoặc tất cả câu
            if ans in ("", "X"):
                logger.warning(
                    f"P1 Q{q} BLANK: ratios={ratios}  "
                    f"max={max(ratios.values()):.3f}  "
                    f"threshold={FILL_THRESHOLD}"
                )
            else:
                logger.debug(f"P1 Q{q}={ans}: ratios={ratios}")

    return answers, details


# ╔════════════════════════════════════════════════════════════════════════╗
# ║   BƯỚC 5b: TRÍCH XUẤT ĐÁP ÁN PHẦN II (8 câu x a/b/c/d x Đ/S)    ║
# ╚════════════════════════════════════════════════════════════════════════╝

def extract_part2(cleaned_img, y_offset=0):
    """
    Đọc 8 câu Đúng/Sai cho mỗi ý a, b, c, d.
    y_offset: bù lệch y do ảnh phồng (từ detect_section_offsets).
    Hybrid 1-pass: max(OpenCV, CNN) cho mỗi bubble.
    Trả về:
      answers: {1: {'a': 'Dung', 'b': 'Sai', ...}, ...}
      details: {1: {'a': {'Dung': 0.7, 'Sai': 0.05}, ...}, ...}
    """
    answers = {}
    details = {}

    for blk in PART2_BLOCKS:
        q = blk["q"]
        sx, sy = blk["start_x"], blk["start_y"]
        q_ans, q_det = {}, {}

        for ri, label in enumerate(PART2_ROWS):
            cy = sy + ri * PART2_STEP_Y + y_offset
            # Cột Đúng
            score_dung, r_dung, _ = _hybrid_score(cleaned_img, sx, cy)
            # Cột Sai
            score_sai, r_sai, _ = _hybrid_score(cleaned_img, sx + PART2_STEP_X, cy)

            q_det[label] = {"Dung": round(score_dung, 3), "Sai": round(score_sai, 3)}

            filled = _detect_filled_choices({"Dung": score_dung, "Sai": score_sai})
            q_ans[label] = _pick_answer(filled)

        answers[q] = q_ans
        details[q] = q_det

    return answers, details


# ╔════════════════════════════════════════════════════════════════════════╗
# ║       BƯỚC 5c: TRÍCH XUẤT ĐÁP ÁN PHẦN III (6 câu điền số)        ║
# ╚════════════════════════════════════════════════════════════════════════╝

def extract_part3(cleaned_img, y_offset=0):
    """
    Đọc 6 câu điền số: dấu trừ (-), dấu phẩy (.), 4 cột số 0-9.
    y_offset: bù lệch y do ảnh phồng (từ detect_section_offsets).
    Mỗi cột chọn digit có fill_ratio cao nhất (nếu > threshold).
    Trả về:
      answers: {1: '1234', 2: '-5.67', ...}
      details: fill ratios chi tiết
    """
    answers = {}
    details = {}

    for blk in PART3_BLOCKS:
        q = blk["q"]
        cols_x = blk["cols_x"]
        q_det = {}

        # 1) Kiểm tra dấu trừ (-)
        sign_score, r_neg, _ = _hybrid_score(
            cleaned_img, blk["sign_x"], PART3_SIGN_Y + y_offset, force_cnn=True
        )
        is_neg = bool(sign_score >= P3_SIGN_SCORE_MIN)
        q_det["sign"] = round(r_neg, 3)

        # 2) Kiểm tra dấu phẩy (.) - cột nào được tô
        #    Comma dùng threshold riêng (0.35) vì dot nhỏ, dễ false positive
        COMMA_THRESHOLD = 0.22
        comma_col = -1
        comma_scores = []
        comma_ratios = []
        for ci, cx in enumerate(cols_x):
            score, ratio, _ = _hybrid_score(
                cleaned_img, cx, PART3_COMMA_Y + y_offset,
                threshold=COMMA_THRESHOLD, force_cnn=True
            )
            comma_scores.append(score)
            comma_ratios.append(round(ratio, 3))
        q_det["comma"] = comma_ratios
        if comma_scores:
            order = sorted(range(len(comma_scores)), key=lambda i: comma_scores[i], reverse=True)
            top_i = order[0]
            top_s = comma_scores[top_i]
            second_s = comma_scores[order[1]] if len(order) > 1 else 0.0
            if top_s >= P3_COMMA_SCORE_MIN and (top_s - second_s) >= P3_COMMA_GAP_MIN:
                comma_col = top_i

        # 3) Đọc 4 cột số (mỗi cột: chọn digit 0-9 có ratio cao nhất)
        #    Dùng logic MAX dominant (giống SBD/MĐ) thay vì _detect_filled_choices
        #    vì 10 bubble/cột khiến adaptive phase quá nhạy → false positive
        digits = []
        digit_det = []
        digit_scores = []
        for ci, cx in enumerate(cols_x):
            col_ratios = {}
            col_scores = {}
            for d in range(10):
                cy = PART3_DIGIT_START_Y + d * PART3_DIGIT_STEP_Y + y_offset
                score, ratio, _ = _hybrid_score(cleaned_img, cx, cy, force_cnn=True)
                col_scores[str(d)] = round(score, 3)
                col_ratios[str(d)] = round(ratio, 3)
            digit_det.append({int(k): v for k, v in col_ratios.items()})
            digit_scores.append({int(k): v for k, v in col_scores.items()})
            # Pick MAX ratio nếu nổi bật (gap > 0.05 so với 2nd)
            sorted_items = sorted(col_scores.items(), key=lambda x: x[1], reverse=True)
            top_d, top_s = sorted_items[0]
            second_s = sorted_items[1][1] if len(sorted_items) > 1 else 0
            if top_s >= P3_DIGIT_SCORE_MIN and (top_s - second_s) >= P3_DIGIT_GAP_MIN:
                digits.append(int(top_d))
            else:
                digits.append(-1)  # Không tô hoặc không rõ
        q_det["digits"] = digit_det
        q_det["digits_score"] = digit_scores

        # 3b) OCR fallback from handwriting boxes (if enabled)
        ocr_digits = []
        ocr_boxes = []
        ocr_ink = []
        digit_count = sum(1 for d in digits if d >= 0)
        allow_ocr = P3_OCR_ENABLE and comma_col >= 0 and digit_count <= 1
        if allow_ocr:
            ocr_digits = [-1] * len(cols_x)
            box_y = PART3_SIGN_Y + y_offset + P3_OCR_BOX_Y_OFFSET
            for ci, cx in enumerate(cols_x):
                box = _find_p3_box_near(cleaned_img, cx, box_y)
                if box is None:
                    box = _default_p3_box(cleaned_img, cx, box_y)
                ocr_boxes.append(box)
                digit, ink_ratio = _ocr_digit_from_box(cleaned_img, box)
                ocr_ink.append(round(ink_ratio, 3))
                if digits[ci] >= 0:
                    continue
                ocr_digits[ci] = digit
                if ocr_digits[ci] >= 0:
                    digits[ci] = ocr_digits[ci]
        q_det["ocr_digits"] = ocr_digits
        q_det["ocr_boxes"] = ocr_boxes
        q_det["ocr_ink"] = ocr_ink
        q_det["picked"] = {
            "sign": is_neg,
            "comma_col": comma_col,
            "digits": digits,
        }
        details[q] = q_det

        # 4) Ghép chuỗi số
        num_str = ""
        for i, d in enumerate(digits):
            if i == comma_col and num_str:
                has_after = any(dd >= 0 for dd in digits[i:])
                if has_after:
                    num_str += "."
            if d >= 0:
                num_str += str(d)
        if is_neg and num_str:
            num_str = "-" + num_str

        answers[q] = num_str

    return answers, details


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                      CHẤM ĐIỂM (Grading)                            ║
# ╚════════════════════════════════════════════════════════════════════════╝

def extract_sbd_made(cleaned_img):
    """
    Đọc Số báo danh (6 chữ số) và Mã đề (3 chữ số).
    Mỗi cột chọn digit có fill_ratio cao nhất (nếu > threshold).
    Trả về: (sbd_str, made_str, details_dict)
    """
    def _read_digit_cols(cols_x, digit_y_list):
        """Đọc cột SBD/MĐ: 10 digits, chọn MAX nổi bật nhất."""
        digits = []
        det = []
        for cx in cols_x:
            col_r = {}
            for d, cy in enumerate(digit_y_list):
                _, r = is_bubble_filled(cleaned_img, cx, cy,
                                        check_circularity=False)
                col_r[d] = round(r, 3)
            det.append(col_r)
            # Pick MAX ratio nếu nổi bật so với 2nd
            sorted_items = sorted(col_r.items(), key=lambda x: x[1], reverse=True)
            top_d, top_r = sorted_items[0]
            second_r = sorted_items[1][1] if len(sorted_items) > 1 else 0
            gap = top_r - second_r
            if top_r >= SBD_MADE_TOP_MIN and gap >= SBD_MADE_GAP_MIN:
                digits.append(top_d)
            elif top_r >= SBD_MADE_FALLBACK_TOP_MIN and gap >= SBD_MADE_FALLBACK_GAP_MIN:
                digits.append(top_d)
            else:
                digits.append(-1)
        return digits, det

    sbd_digits, sbd_det = _read_digit_cols(SBD_COLS_X, SBD_MADE_DIGIT_Y)
    made_digits, made_det = _read_digit_cols(MADE_COLS_X, SBD_MADE_DIGIT_Y)

    sbd_str = "".join(str(d) if d >= 0 else "?" for d in sbd_digits)
    made_str = "".join(str(d) if d >= 0 else "?" for d in made_digits)

    details = {"sbd": sbd_det, "made": made_det}
    return sbd_str, made_str, details


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                      CHẤM ĐIỂM (Grading)                            ║
# ╚════════════════════════════════════════════════════════════════════════╝

def grade_part1(student, correct):
    """So sánh Part I. Trả về (score, results_dict)."""
    results = {}
    score = 0
    for q in range(1, 41):
        s = student.get(q, "")
        c = correct.get(q, "")
        ok = (s == c and s not in ("", "X"))
        if ok:
            score += 1
        results[q] = {"student": s, "correct": c, "is_correct": ok}
    return score, results


def grade_part2(student, correct):
    """So sánh Part II. Trả về (score, results_dict). 1 điểm nếu đúng cả 4 ý."""
    results = {}
    score = 0
    for q in range(1, 9):
        q_res = {}
        n_correct = 0
        for row in PART2_ROWS:
            s = student.get(q, {}).get(row, "")
            c = correct.get(q, {}).get(row, "")
            ok = (s == c and s not in ("", "X"))
            if ok:
                n_correct += 1
            q_res[row] = {"student": s, "correct": c, "is_correct": ok}
        if n_correct == 4:
            score += 1
        results[q] = q_res
    return score, results


def grade_part3(student, correct):
    """So sánh Part III. Trả về (score, results_dict)."""
    results = {}
    score = 0
    for q in range(1, 7):
        s = student.get(q, "")
        c = correct.get(q, "")
        ok = (s == c and s != "")
        if ok:
            score += 1
        results[q] = {"student": s, "correct": c, "is_correct": ok}
    return score, results


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                  VẼ KẾT QUẢ LÊN ẢNH (Visualization)                ║
# ╚════════════════════════════════════════════════════════════════════════╝

# Màu sắc (BGR) — đậm, nổi bật
COLOR_CORRECT   = (0, 180, 0)     # Xanh lá đậm  - đáp án HS đúng
COLOR_WRONG     = (0, 0, 220)     # Đỏ đậm       - đáp án HS sai
COLOR_RIGHT_ANS = (0, 180, 0)     # Xanh lá đậm - đáp án đúng (khi HS sai)
COLOR_UNANSWERED = (0, 200, 255)  # Vàng đậm - chưa khoanh / bỏ trống
THICKNESS_MARK  = 5


def draw_results_part1(image, results, y_offset=0):
    """Vẽ vòng tròn kết quả cho Part I."""
    for q, res in results.items():
        col_idx = (q - 1) // 10
        row_idx = (q - 1) % 10
        if col_idx >= len(PART1_COLS):
            continue
        cfg = PART1_COLS[col_idx]
        cy = int(cfg["start_y"] + row_idx * cfg["step_y"] + y_offset)
        student_ans = res["student"]
        is_blank = student_ans in ("", "-", "X")

        for ci, choice in enumerate(PART1_CHOICES):
            cx = int(cfg["start_x"] + ci * cfg["step_x"])
            # Đánh dấu đáp án học sinh
            if choice == student_ans:
                color = COLOR_CORRECT if res["is_correct"] else COLOR_WRONG
                cv2.circle(image, (cx, cy), BUBBLE_RADIUS + 3, color, THICKNESS_MARK)
            # Đánh dấu đáp án đúng nếu HS sai
            if choice == res["correct"] and not res["is_correct"]:
                cv2.circle(image, (cx, cy), BUBBLE_RADIUS + 5, COLOR_RIGHT_ANS, THICKNESS_MARK)
            # Vàng: chưa khoanh → highlight đáp án đúng bằng màu vàng
            if is_blank and choice == res["correct"]:
                cv2.circle(image, (cx, cy), BUBBLE_RADIUS + 5, COLOR_UNANSWERED, THICKNESS_MARK)
    return image


def draw_results_part2(image, results, y_offset=0):
    """Vẽ vòng tròn kết quả cho Part II."""
    for q, q_res in results.items():
        blk = PART2_BLOCKS[q - 1]
        sx, sy = blk["start_x"], blk["start_y"]
        for ri, label in enumerate(PART2_ROWS):
            res = q_res[label]
            student_ans = res["student"]
            is_blank = student_ans in ("", "X")
            cy = int(sy + ri * PART2_STEP_Y + y_offset)
            for ci, col_name in enumerate(["Dung", "Sai"]):
                cx = int(sx + ci * PART2_STEP_X)
                if col_name == student_ans:
                    color = COLOR_CORRECT if res["is_correct"] else COLOR_WRONG
                    cv2.circle(image, (cx, cy), BUBBLE_RADIUS + 3, color, THICKNESS_MARK)
                if col_name == res["correct"] and not res["is_correct"]:
                    cv2.circle(image, (cx, cy), BUBBLE_RADIUS + 5, COLOR_RIGHT_ANS, THICKNESS_MARK)
                # Vàng: chưa khoanh → highlight đáp án đúng
                if is_blank and col_name == res["correct"]:
                    cv2.circle(image, (cx, cy), BUBBLE_RADIUS + 5, COLOR_UNANSWERED, THICKNESS_MARK)
    return image


def _parse_p3_string(s, num_cols=4):
    """Parse Part III answer string → (is_neg, comma_col, digits[4])."""
    if s is not None and not isinstance(s, str):
        s = str(s)
    if not s or s == '-':
        return False, -1, [-1] * num_cols
    is_neg = s.startswith('-')
    s = s.lstrip('-')
    comma_col = -1
    if '.' in s:
        comma_col = s.index('.')
        s = s.replace('.', '')
    digits = [-1] * num_cols
    for i, ch in enumerate(s):
        if i < num_cols:
            digits[i] = int(ch)
    return is_neg, comma_col, digits


def draw_results_part3(image, results, student_details, y_offset=0):
    """Vẽ kết quả Part III: khoanh đỏ bubble sai, xanh bubble đúng."""
    for q, res in results.items():
        blk = PART3_BLOCKS[q - 1]
        q_det = student_details.get(q, {})
        is_correct = res["is_correct"]
        mark_color = COLOR_CORRECT if is_correct else COLOR_WRONG

        # --- Khoanh bubble học sinh đã tô ---
        picked = q_det.get("picked")
        if picked:
            if picked.get("sign"):
                cv2.circle(image, (int(blk["sign_x"]), PART3_SIGN_Y + y_offset),
                           BUBBLE_RADIUS + 3, mark_color, THICKNESS_MARK)
            comma_col = picked.get("comma_col", -1)
            if 0 <= comma_col < len(blk["cols_x"]):
                cx = int(blk["cols_x"][comma_col])
                cv2.circle(image, (cx, PART3_COMMA_Y + y_offset),
                           BUBBLE_RADIUS + 3, mark_color, THICKNESS_MARK)
            for ci, d in enumerate(picked.get("digits", [])):
                if 0 <= d <= 9 and ci < len(blk["cols_x"]):
                    cx = int(blk["cols_x"][ci])
                    cy = int(PART3_DIGIT_START_Y + d * PART3_DIGIT_STEP_Y + y_offset)
                    cv2.circle(image, (cx, cy),
                               BUBBLE_RADIUS + 3, mark_color, THICKNESS_MARK)

        # OCR boxes (handwriting) if present
        ocr_boxes = q_det.get("ocr_boxes", [])
        ocr_digits = q_det.get("ocr_digits", [])
        for i, box in enumerate(ocr_boxes):
            if box is None or i >= len(ocr_digits):
                continue
            if ocr_digits[i] < 0:
                continue
            x1, y1, x2, y2 = box
            cv2.rectangle(image, (x1, y1), (x2, y2), (255, 120, 0), 2)
        else:
            # Sign
            sign_r = q_det.get("sign", 0)
            if sign_r > FILL_THRESHOLD:
                cv2.circle(image, (int(blk["sign_x"]), PART3_SIGN_Y + y_offset),
                           BUBBLE_RADIUS + 3, mark_color, THICKNESS_MARK)
            # Comma
            for ci, r in enumerate(q_det.get("comma", [])):
                if r > FILL_THRESHOLD and ci < len(blk["cols_x"]):
                    cx = int(blk["cols_x"][ci])
                    cv2.circle(image, (cx, PART3_COMMA_Y + y_offset),
                               BUBBLE_RADIUS + 3, mark_color, THICKNESS_MARK)
            # Digits
            for ci, col_ratios in enumerate(q_det.get("digits", [])):
                if ci >= len(blk["cols_x"]):
                    continue
                cx = int(blk["cols_x"][ci])
                for d, r in col_ratios.items():
                    if r > FILL_THRESHOLD:
                        cy = int(PART3_DIGIT_START_Y + int(d) * PART3_DIGIT_STEP_Y + y_offset)
                        cv2.circle(image, (cx, cy),
                                   BUBBLE_RADIUS + 3, mark_color, THICKNESS_MARK)

        # --- Nếu sai: khoanh XANH đáp án đúng ---
        if not is_correct and res["correct"]:
            c_neg, c_comma, c_digits = _parse_p3_string(res["correct"])
            # Sign đúng
            if c_neg:
                cv2.circle(image, (int(blk["sign_x"]), PART3_SIGN_Y + y_offset),
                           BUBBLE_RADIUS + 5, COLOR_RIGHT_ANS, THICKNESS_MARK)
            # Comma đúng
            if 0 <= c_comma < len(blk["cols_x"]):
                cx = int(blk["cols_x"][c_comma])
                cv2.circle(image, (cx, PART3_COMMA_Y + y_offset),
                           BUBBLE_RADIUS + 5, COLOR_RIGHT_ANS, THICKNESS_MARK)
            # Digits đúng
            for ci, d in enumerate(c_digits):
                if 0 <= d <= 9 and ci < len(blk["cols_x"]):
                    cx = int(blk["cols_x"][ci])
                    cy = int(PART3_DIGIT_START_Y + d * PART3_DIGIT_STEP_Y + y_offset)
                    cv2.circle(image, (cx, cy),
                               BUBBLE_RADIUS + 5, COLOR_RIGHT_ANS, THICKNESS_MARK)

        # --- Text đáp án ---
        first_cx = blk["cols_x"][0]
        text_y = PART3_SIGN_Y - 25 + y_offset
        cv2.putText(image, f"={res['student']}",
                    (int(first_cx - 10), int(text_y)),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, mark_color, 2)
        if not is_correct:
            cv2.putText(image, f"({res['correct']})",
                        (int(first_cx - 10), int(text_y - 18)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.38, COLOR_RIGHT_ANS, 2)
    return image


# ╔════════════════════════════════════════════════════════════════════════╗
# ║              CALIBRATION: VẼ LƯỚI BUBBLE DỰ KIẾN                    ║
# ╚════════════════════════════════════════════════════════════════════════╝

def draw_bubble_grid(warped_image, offsets=None):
    """Vẽ tất cả vị trí bubble dự kiến lên ảnh warped (debug/calibration).
    offsets: dict {"part1": dy, "part2": dy, "part3": dy} — bù lệch y cục bộ.
    """
    vis = warped_image.copy()
    dy1 = offsets.get("part1", 0) if offsets else 0
    dy2 = offsets.get("part2", 0) if offsets else 0
    dy3 = offsets.get("part3", 0) if offsets else 0

    # Part I: vòng tròn xanh lá
    for cfg in PART1_COLS:
        col_rows = cfg.get("num_rows", PART1_NUM_ROWS)
        for row in range(col_rows):
            cy = int(cfg["start_y"] + row * cfg["step_y"] + dy1)
            for ci in range(4):
                cx = int(cfg["start_x"] + ci * cfg["step_x"])
                cv2.circle(vis, (cx, cy), BUBBLE_RADIUS, (0, 255, 0), 1)

    # Part II: vòng tròn xanh dương
    for blk in PART2_BLOCKS:
        for ri in range(4):
            cy = int(blk["start_y"] + ri * PART2_STEP_Y + dy2)
            for ci in range(2):
                cx = int(blk["start_x"] + ci * PART2_STEP_X)
                cv2.circle(vis, (cx, cy), BUBBLE_RADIUS, (255, 100, 0), 1)

    # SBD + Mã đề: ô vuông + vòng tròn bên trong
    r = BUBBLE_RADIUS
    cell_half = r + 4  # nửa cạnh ô vuông bao quanh bubble

    for label, cols_x in [("SBD", SBD_COLS_X), ("MD", MADE_COLS_X)]:
        for ci, cx in enumerate(cols_x):
            cx_i = int(cx)
            # Ô vuông header (ô ghi số bên trên, y ~ 130-155)
            hdr_y = SBD_MADE_DIGIT_Y[0] - 40
            cv2.rectangle(vis, (cx_i - cell_half, hdr_y - cell_half),
                          (cx_i + cell_half, hdr_y + cell_half), (200, 0, 200), 1)

            for d, cy in enumerate(SBD_MADE_DIGIT_Y):
                cy_i = int(cy)
                # Ô vuông bao quanh mỗi cell
                cv2.rectangle(vis, (cx_i - cell_half, cy_i - cell_half),
                              (cx_i + cell_half, cy_i + cell_half), (200, 0, 200), 1)
                # Vòng tròn bubble bên trong
                cv2.circle(vis, (cx_i, cy_i), r, (0, 200, 200), 1)

            # Nhãn cột (SBD1..6, MD1..3)
            cv2.putText(vis, f"{ci}", (cx_i - 4, hdr_y + 4),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.35, (200, 0, 200), 1)

    # Part III: vòng tròn đỏ
    for blk in PART3_BLOCKS:
        # Dấu trừ
        cv2.circle(vis, (int(blk["sign_x"]), PART3_SIGN_Y + dy3), BUBBLE_RADIUS, (0, 0, 255), 1)
        for cx in blk["cols_x"]:
            # Dấu phẩy
            cv2.circle(vis, (int(cx), PART3_COMMA_Y + dy3), BUBBLE_RADIUS, (0, 0, 255), 1)
            # Số 0-9
            for d in range(10):
                cy = int(PART3_DIGIT_START_Y + d * PART3_DIGIT_STEP_Y + dy3)
                cv2.circle(vis, (int(cx), cy), BUBBLE_RADIUS, (0, 0, 255), 1)

    # Vẽ vùng erase mask (vàng mờ = vùng xóa, lỗ tròn = bubble bảo vệ)
    overlay = vis.copy()
    for (x1, y1, x2, y2) in TEXT_ERASE_REGIONS:
        cv2.rectangle(overlay, (max(0, int(x1)), max(0, int(y1))),
                       (min(WARP_WIDTH, int(x2)), min(WARP_HEIGHT, int(y2))),
                       (0, 200, 255), -1)
    # Đục lỗ tròn (hiển thị vùng bảo vệ bubble)
    for (cx, cy) in ALL_BUBBLE_CENTERS:
        cv2.circle(overlay, (cx, cy), BUBBLE_PROTECT_RADIUS, vis[cy, cx].tolist() if 0 <= cy < WARP_HEIGHT and 0 <= cx < WARP_WIDTH else (255,255,255), -1)
    cv2.addWeighted(overlay, 0.25, vis, 0.75, 0, vis)

    return vis


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                     PIPELINE CHÍNH (Main)                            ║
# ╚════════════════════════════════════════════════════════════════════════╝

def process_sheet(image_path, correct_answers=None, debug=False, pre_warped=False, provided_corners=None):
    """
    Pipeline đầy đủ: phát hiện góc → warp → tiền xử lý → đọc đáp án → chấm điểm.

    Tham số:
      image_path      : đường dẫn ảnh phiếu
      correct_answers : dict với keys 'part1', 'part2', 'part3'
        part1: {1: 'A', 2: 'B', ...}
        part2: {1: {'a': 'Dung', 'b': 'Sai', ...}, ...}
        part3: {1: '1234', 2: '-5.67', ...}
      debug           : True → lưu ảnh calibration + threshold
      pre_warped      : True → bỏ qua detect corner (ảnh đã thẳng)
      provided_corners: Tọa độ 4 góc được truyền từ frontend (nếu có)
    """
    print(f"\n{'='*60}")
    print(f"  Xử lý: {os.path.basename(image_path)}")
    print(f"{'='*60}")

    # --- Load ảnh ---
    base = os.path.splitext(image_path)[0]
    image = cv2.imread(image_path)
    if image is None:
        print(f"[LỖI] Không đọc được ảnh: {image_path}")
        return None

    # --- Fix EXIF orientation (ảnh camera điện thoại) ---
    try:
        pil_img = Image.open(image_path)
        exif = pil_img._getexif()
        if exif:
            orientation_key = next(
                (k for k, v in ExifTags.TAGS.items() if v == 'Orientation'), None
            )
            if orientation_key and orientation_key in exif:
                orient = exif[orientation_key]
                if orient == 3:
                    image = cv2.rotate(image, cv2.ROTATE_180)
                    print("[EXIF] Xoay 180°")
                elif orient == 6:
                    image = cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
                    print("[EXIF] Xoay 90° (phải)")
                elif orient == 8:
                    image = cv2.rotate(image, cv2.ROTATE_90_COUNTERCLOCKWISE)
                    print("[EXIF] Xoay 90° (trái)")
    except Exception:
        pass  # Không có EXIF hoặc lỗi → bỏ qua

    # --- Bước 1-2: Detect corners + Warp ---
    if pre_warped:
        warped = cv2.resize(image, (WARP_WIDTH, WARP_HEIGHT))
        method = "pre_warped"
        _orig_method = method
        corners = None
        all_candidates = []
        print(f"[OK] Ảnh pre-warped ({WARP_WIDTH}x{WARP_HEIGHT})")
    elif provided_corners is not None and len(provided_corners) == 4:
        try:
            pts = np.array(provided_corners, dtype="float32")
            ordered_pts = order_points(pts)
            warped = _warp_to_rect(image, ordered_pts)
            method = "frontend_corners"
            _orig_method = method
            corners = ordered_pts
            all_candidates = []
            print(f"[OK] Sử dụng tọa độ góc từ Frontend: {ordered_pts.astype(int).tolist()}")
        except Exception as e:
            print(f"[LỖI] parse frontend_corners: {e}")
            return None
    else:
        try:
            detect_result = detect_paper_and_warp(image, debug=debug)
            warped = detect_result["warped"]
            method = detect_result["method"]
            _orig_method = method
            corners = detect_result["corners"]
            all_candidates = detect_result.get("_candidates", [])
            print(f"[OK] Phát hiện bằng: {method}")
            print(f"[OK] 4 góc: {corners.astype(int).tolist()}")
            print(f"[OK] Warped → {WARP_WIDTH}x{WARP_HEIGHT}")
            # Lưu ảnh debug detection nếu có
            if debug and detect_result["debug_image"] is not None:
                dbg_path = f"{base}_detect.jpg"
                cv2.imwrite(dbg_path, detect_result["debug_image"])
                print(f"[DEBUG] Ảnh detect → {dbg_path}")
        except ValueError as e:
            print(f"[LỖI] {e}")
            return None

    # --- Bước 3: Tiền xử lý ---
    # Pipeline: erase_printed_text → auto-detect phone → illumination norm + CLAHE
    gray, thresh, cleaned = preprocess(warped)
    print("[OK] Grayscale detection (pencil-friendly)")

    # --- Bước 3b: Detect marker offsets (bù ảnh phồng) ---
    offsets = detect_section_offsets(gray)
    any_offset = any(v != 0 for v in offsets.values())
    if any_offset:
        print(f"[OK] Marker offsets: P1={offsets['part1']:+d}px  P2={offsets['part2']:+d}px  P3={offsets['part3']:+d}px")
    else:
        print("[OK] Marker offsets: 0 (ảnh thẳng)")

    p3_offset = detect_part3_offset_from_digits(gray)
    if p3_offset is not None:
        offsets["part3"] = p3_offset
        print(f"[OK] Part3 offset (digits): {p3_offset:+d}px")

    # --- Debug output (SAU khi có offsets để calibration chính xác) ---
    if debug:
        cv2.imwrite(f"{base}_calibration.jpg", draw_bubble_grid(warped, offsets=offsets))
        cv2.imwrite(f"{base}_gray.jpg", gray)
        cv2.imwrite(f"{base}_thresh.jpg", thresh)
        cv2.imwrite(f"{base}_cleaned.jpg", cleaned)
        print(f"[DEBUG] Đã lưu: _calibration.jpg, _gray.jpg, _thresh.jpg, _cleaned.jpg")

    # --- Bước 4-5: Đọc đáp án (dùng ảnh GRAYSCALE, không binary) ---
    sbd, made, sbd_det = extract_sbd_made(gray)
    p1_ans, p1_det = extract_part1(gray, y_offset=offsets["part1"])
    p2_ans, p2_det = extract_part2(gray, y_offset=offsets["part2"])
    p3_ans, p3_det = extract_part3(gray, y_offset=offsets["part3"])

    # --- Retry với method khác nếu SBD hoặc Mã đề có '?' (VÔ HIỆU HÓA vì SBD/Made không bắt buộc) ---
    # sbd_has_q = '?' in str(sbd)
    # made_has_q = '?' in str(made)
    # if (sbd_has_q or made_has_q) and len(all_candidates) > 1:
    #     print(f"\n  [RETRY] SBD='{sbd}' MĐ='{made}' có '?' → thử method khác (đã tắt).")


    # Cảnh báo nếu sau tất cả vẫn còn '?'
    if '?' in str(sbd) or '?' in str(made):
        print(f"\n  [CẢNH BÁO] SBD='{sbd}' MĐ='{made}' — "
              f"học sinh chưa tô hoặc ảnh không rõ. Đã thử {len(all_candidates)} method.")

    # --- Ghi lại debug images nếu retry đã chuyển method ---
    if debug and not pre_warped and method != _orig_method:
        cv2.imwrite(f"{base}_calibration.jpg", draw_bubble_grid(warped, offsets=offsets))
        cv2.imwrite(f"{base}_gray.jpg", gray)
        cv2.imwrite(f"{base}_thresh.jpg", thresh)
        cv2.imwrite(f"{base}_cleaned.jpg", cleaned)
        print(f"[DEBUG] Ghi lại debug images cho method: {method}")

    # --- In kết quả ---
    print(f"\n  SỐ BÁO DANH: {sbd}")
    print(f"  MÃ ĐỀ      : {made}")
    _print_answers(p1_ans, p2_ans, p3_ans)

    # --- Bước 6: Chấm điểm (nếu có đáp án đúng) ---
    # Vẽ lên mask trống → blend vào warped (addWeighted overlay)
    result_mask = np.zeros_like(warped)   # mask trống để vẽ kết quả
    result_image = warped.copy()
    total_score = None

    if correct_answers:
        scores = {}
        if "part1" in correct_answers:
            s, r = grade_part1(p1_ans, correct_answers["part1"])
            scores["part1"] = s
            draw_results_part1(result_mask, r, y_offset=offsets["part1"])
            print(f"\n  Phần I  : {s}/40")

        if "part2" in correct_answers:
            s, r = grade_part2(p2_ans, correct_answers["part2"])
            scores["part2"] = s
            draw_results_part2(result_mask, r, y_offset=offsets["part2"])
            print(f"  Phần II : {s}/8")

        if "part3" in correct_answers:
            s, r = grade_part3(p3_ans, correct_answers["part3"])
            scores["part3"] = s
            draw_results_part3(result_mask, r, p3_det, y_offset=offsets["part3"])
            print(f"  Phần III: {s}/6")

        total_score = sum(scores.values())
        max_score = sum([40 if "part1" in scores else 0,
                         8 if "part2" in scores else 0,
                         6 if "part3" in scores else 0])
        print(f"  ─────────────────")
        print(f"  TỔNG ĐIỂM: {total_score}/{max_score}")

        # Ghi điểm to lên mask
        cv2.putText(result_mask, f"DIEM: {total_score}/{max_score}",
                    (30, 55), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0, 0, 220), 3)

    # Ghi SBD + Mã đề lên mask
    cv2.putText(result_mask, f"SBD: {sbd}  MD: {made}",
                (30, 95), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (180, 0, 0), 2)

    # Blend mask lên warped: nền bài + kết quả bán trong suốt
    result_image = cv2.addWeighted(result_image, 1, result_mask, 0.7, 0)

    # --- Lưu ảnh kết quả (warped) ---
    out_path = f"{base}_result.jpg"
    cv2.imwrite(out_path, result_image)
    print(f"\n[OK] Ảnh kết quả: {out_path}")

    # --- Inverse Warp: nắn kết quả ngược về ảnh gốc ---
    if not pre_warped and corners is not None:
        try:
            dst_pts = np.array([
                [0, 0], [WARP_WIDTH - 1, 0],
                [WARP_WIDTH - 1, WARP_HEIGHT - 1], [0, WARP_HEIGHT - 1]
            ], dtype="float32")
            inv_M = cv2.getPerspectiveTransform(dst_pts, corners)
            h_orig, w_orig = image.shape[:2]
            # Nắn mask kết quả ngược về perspective gốc
            inv_mask = cv2.warpPerspective(result_mask, inv_M, (w_orig, h_orig))
            # Overlay lên ảnh gốc
            overlay_image = cv2.addWeighted(image, 1, inv_mask, 0.8, 0)
            overlay_path = f"{base}_overlay.jpg"
            cv2.imwrite(overlay_path, overlay_image)
            print(f"[OK] Ảnh overlay (inverse warp): {overlay_path}")
        except Exception as e:
            print(f"[WARN] Inverse warp failed: {e}")

    # --- Crop vùng tên học sinh từ ảnh warped gốc (sạch) ---
    name_path = ""
    nx, ny, nw, nh = NAME_REGION
    if nw > 0 and nh > 0:
        h_img, w_img = warped.shape[:2]
        y1 = max(0, ny)
        y2 = min(h_img, ny + nh)
        x1 = max(0, nx)
        x2 = min(w_img, nx + nw)
        if y2 > y1 and x2 > x1:
            name_crop = warped[y1:y2, x1:x2]
            name_path = f"{base}_name.jpg"
            cv2.imwrite(name_path, name_crop)
            print(f"[OK] Ảnh tên: {name_path}")

    _load_bubble_cnn()
    cnn_status = "ready" if _CNN_READY else f"error:{_CNN_ERROR}"

    return {
        "sbd": sbd, "made": made,
        "part1": p1_ans, "part2": p2_ans, "part3": p3_ans,
        "score": total_score,
        "max_score": max_score if correct_answers else None,
        "scores": scores if correct_answers else {},
        "details": {"sbd": sbd_det, "part1": p1_det, "part2": p2_det, "part3": p3_det},
        "name_image_path": name_path,
        "detect_method": method if not pre_warped else "pre_warped",
        "offsets": offsets,
        "cnn_status": cnn_status,
    }


def _print_answers(p1, p2, p3):
    """In đáp án học sinh ra console."""
    print("\n  ── Đáp án học sinh ──")
    print("  Phần I:")
    for q in range(1, 41):
        a = p1.get(q, "")
        flag = " [!]" if a in ("X", "") else ""
        end = "  " if q % 5 != 0 else "\n"
        print(f"    Q{q:2d}: {a or '-'}{flag}", end=end)
    if 40 % 5 != 0:
        print()

    print("  Phần II:")
    for q in range(1, 9):
        qa = p2.get(q, {})
        parts = [f"{r}={qa.get(r, '-')}" for r in PART2_ROWS]
        print(f"    Q{q}: {', '.join(parts)}")

    print("  Phần III:")
    for q in range(1, 7):
        print(f"    Q{q}: {p3.get(q, '-') or '-'}")


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                      XỬ LÝ HÀNG LOẠT                                ║
# ╚════════════════════════════════════════════════════════════════════════╝

def batch_process(folder, correct_answers=None, debug=False,
                   results_dir="results", excel_path="results/bang_diem.xlsx"):
    """
    Chấm tất cả ảnh .jpg/.png trong thư mục.
    Tự động lưu JSON từng bài + xuất Excel tổng hợp.
    """
    exts = (".jpg", ".jpeg", ".png", ".bmp")
    skip = ("_result", "_calibration", "_thresh", "_cleaned",
            "_simtest", "_sim59", "_simdup")
    files = [f for f in os.listdir(folder)
             if f.lower().endswith(exts) and not any(s in f for s in skip)]
    print(f"Tìm thấy {len(files)} ảnh trong {folder}")
    results = {}
    for fname in sorted(files):
        path = os.path.join(folder, fname)
        res = process_sheet(path, correct_answers, debug=debug)
        results[fname] = res
        if res:
            save_result(res, results_dir, correct_answers)

    # Xuất Excel tổng hợp
    valid = [r for r in results.values() if r is not None]
    if valid:
        export_excel(valid, excel_path)

    return results


# ╔════════════════════════════════════════════════════════════════════════╗
# ║          ĐÁP ÁN MẪU (thay đổi theo đề thi thực tế)                 ║
# ╚════════════════════════════════════════════════════════════════════════╝

SAMPLE_CORRECT = {
    "part1": {i: ["A", "B", "C", "D"][(i - 1) % 4] for i in range(1, 41)},
    "part2": {q: {"a": "Dung", "b": "Sai", "c": "Dung", "d": "Sai"} for q in range(1, 9)},
    "part3": {1: "1234", 2: "-5.67", 3: "0042", 4: "8", 5: "-9999", 6: "31.41"},
}


# ╔════════════════════════════════════════════════════════════════════════╗
# ║      TEST: Tạo phiếu mô phỏng (tô bubble lên ảnh trắng)           ║
# ╚════════════════════════════════════════════════════════════════════════╝

def create_test_sheet(blank_path, output_path, ans_p1=None, ans_p2=None, ans_p3=None,
                      sbd_str=None, made_str=None):
    """
    Tô bubble lên phiếu trắng để tạo ảnh test.
    ans_p1: {1: 'A', 2: 'C', ...}
    ans_p2: {1: {'a': 'Dung', 'b': 'Sai', ...}, ...}
    ans_p3: {1: {'sign': False, 'comma_col': -1, 'digits': [1,2,3,4]}, ...}
    sbd_str: '002568' (6 chữ số)
    made_str: '001' (3 chữ số)
    """
    image = cv2.imread(blank_path)
    if image is None:
        return None
    corners = detect_corners(image)
    warped = warp_perspective(image, corners)

    r = BUBBLE_RADIUS - 2  # Tô nhỏ hơn bubble thật một chút

    # Tô Part I (hỗ trợ tô trùng: "AC" → tô cả A và C)
    if ans_p1:
        choice_map = {"A": 0, "B": 1, "C": 2, "D": 3}
        for q, ans in ans_p1.items():
            col_idx, row_idx = (q - 1) // 10, (q - 1) % 10
            cfg = PART1_COLS[col_idx]
            cy = int(cfg["start_y"] + row_idx * cfg["step_y"])
            for ch in str(ans):
                if ch in choice_map:
                    ci = choice_map[ch]
                    cx = int(cfg["start_x"] + ci * cfg["step_x"])
                    cv2.circle(warped, (cx, cy), r, (0, 0, 0), -1)

    # Tô Part II (hỗ trợ tô trùng: "X" → tô cả Đúng và Sai)
    if ans_p2:
        col_map = {"Dung": 0, "Sai": 1}
        for q, q_ans in ans_p2.items():
            blk = PART2_BLOCKS[q - 1]
            for ri, label in enumerate(PART2_ROWS):
                a = q_ans.get(label, "")
                if a == "X":
                    cy = int(blk["start_y"] + ri * PART2_STEP_Y)
                    for ci in range(2):
                        cx = int(blk["start_x"] + ci * PART2_STEP_X)
                        cv2.circle(warped, (cx, cy), r, (0, 0, 0), -1)
                elif a in col_map:
                    cx = int(blk["start_x"] + col_map[a] * PART2_STEP_X)
                    cy = int(blk["start_y"] + ri * PART2_STEP_Y)
                    cv2.circle(warped, (cx, cy), r, (0, 0, 0), -1)

    # Tô Part III
    if ans_p3:
        for q, data in ans_p3.items():
            blk = PART3_BLOCKS[q - 1]
            if data.get("sign"):
                cv2.circle(warped, (int(blk["sign_x"]), PART3_SIGN_Y), r, (0, 0, 0), -1)
            # comma_col: int hoặc list[int] (tô trùng)
            cc = data.get("comma_col", -1)
            ccs = cc if isinstance(cc, list) else [cc]
            for c in ccs:
                if 0 <= c < len(blk["cols_x"]):
                    cv2.circle(warped, (int(blk["cols_x"][c]), PART3_COMMA_Y), r, (0, 0, 0), -1)
            # digits: mỗi phần tử là int hoặc list[int] (tô trùng)
            for ci, d in enumerate(data.get("digits", [])):
                ds = d if isinstance(d, list) else [d]
                for dd in ds:
                    if 0 <= dd <= 9 and ci < len(blk["cols_x"]):
                        cx = int(blk["cols_x"][ci])
                        cy = int(PART3_DIGIT_START_Y + dd * PART3_DIGIT_STEP_Y)
                        cv2.circle(warped, (cx, cy), r, (0, 0, 0), -1)

    # Tô SBD
    if sbd_str:
        for ci, ch in enumerate(sbd_str):
            if ci < len(SBD_COLS_X) and ch.isdigit():
                d = int(ch)
                cx = SBD_COLS_X[ci]
                cy = SBD_MADE_DIGIT_Y[d]
                cv2.circle(warped, (int(cx), int(cy)), r, (0, 0, 0), -1)

    # Tô Mã đề
    if made_str:
        for ci, ch in enumerate(made_str):
            if ci < len(MADE_COLS_X) and ch.isdigit():
                d = int(ch)
                cx = MADE_COLS_X[ci]
                cy = SBD_MADE_DIGIT_Y[d]
                cv2.circle(warped, (int(cx), int(cy)), r, (0, 0, 0), -1)

    cv2.imwrite(output_path, warped)
    print(f"[TEST] Phiếu mô phỏng: {output_path}")
    return output_path


# ╔════════════════════════════════════════════════════════════════════════╗
# ║                          ENTRY POINT (CLI)                            ║
# ╚════════════════════════════════════════════════════════════════════════╝

if __name__ == "__main__":
    import sys

    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    ROOT_DIR   = os.path.dirname(SCRIPT_DIR)

    # Load template + đáp án
    template_path = os.path.join(ROOT_DIR, "templates", "template_default.json")
    answers_path  = os.path.join(ROOT_DIR, "answers", "de_mau.json")
    results_dir   = os.path.join(SCRIPT_DIR, "results")
    os.makedirs(results_dir, exist_ok=True)

    if os.path.exists(template_path):
        load_template(template_path)

    correct = None
    if os.path.exists(answers_path):
        correct = load_answers(answers_path)

    # Chấm ảnh từ command line
    if len(sys.argv) < 2:
        print("Cách dùng: python hi.py <ảnh1.jpg> [ảnh2.jpg] ...")
        print("Hoặc dùng giao diện web: streamlit run app.py")
        sys.exit(0)

    all_results = []
    for img_path in sys.argv[1:]:
        if os.path.isfile(img_path):
            result = process_sheet(img_path, correct_answers=correct, debug=True)
            if result:
                save_result(result, results_dir, correct)
                all_results.append(result)

    if len(all_results) > 1:
        excel_path = os.path.join(results_dir, "bang_diem.xlsx")
        export_excel(all_results, excel_path)

    print(f"\nXong — {len(all_results)} bài đã chấm.")
